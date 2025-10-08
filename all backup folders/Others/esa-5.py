import sys
import os
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QLabel, QVBoxLayout, QWidget,
    QFileDialog, QMessageBox, QTabWidget, QTextEdit, QComboBox, QHBoxLayout
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QTextCursor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelComparator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel File Comparator")
        self.setGeometry(200, 200, 900, 700)

        # Layouts
        layout = QVBoxLayout()

        # File selection
        self.file1_label = QLabel("File 1: Not selected")
        self.file2_label = QLabel("File 2: Not selected")

        self.select_file1_button = QPushButton("Select First Excel File")
        self.select_file2_button = QPushButton("Select Second Excel File")
        self.select_file1_button.clicked.connect(self.load_file1)
        self.select_file2_button.clicked.connect(self.load_file2)

        layout.addWidget(self.file1_label)
        layout.addWidget(self.select_file1_button)
        layout.addWidget(self.file2_label)
        layout.addWidget(self.select_file2_button)

        # Column selectors
        self.col1_selector = QComboBox()
        self.col2_selector = QComboBox()
        col_layout = QHBoxLayout()
        col_layout.addWidget(QLabel("Compare column from File 1:"))
        col_layout.addWidget(self.col1_selector)
        col_layout.addWidget(QLabel("Compare column from File 2:"))
        col_layout.addWidget(self.col2_selector)
        layout.addLayout(col_layout)

        # Compare button
        self.compare_button = QPushButton("Compare Files")
        self.compare_button.clicked.connect(self.compare_columns)
        layout.addWidget(self.compare_button)

        # Tabs for results
        self.tabs = QTabWidget()
        self.only_in_file1_text = QTextEdit()
        self.only_in_file2_text = QTextEdit()
        self.common_text = QTextEdit()
        self.summary_text = QTextEdit()
        self.merged_text = QTextEdit()

        self.only_in_file1_text.setReadOnly(True)
        self.only_in_file2_text.setReadOnly(True)
        self.common_text.setReadOnly(True)
        self.summary_text.setReadOnly(True)
        self.merged_text.setReadOnly(True)

        self.tabs.addTab(self.only_in_file1_text, "Only in First File")
        self.tabs.addTab(self.only_in_file2_text, "Only in Second File")
        self.tabs.addTab(self.common_text, "Common Values")
        self.tabs.addTab(self.summary_text, "Summary")
        self.tabs.addTab(self.merged_text, "Merged Results")

        layout.addWidget(self.tabs)

        # Export button
        self.export_button = QPushButton("Export Results to Excel")
        self.export_button.clicked.connect(self.export_results)
        layout.addWidget(self.export_button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # Data storage
        self.df1 = None
        self.df2 = None
        self.results = None

    def load_file1(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open First Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                self.df1 = pd.read_excel(file_path)
                self.file1_label.setText(f"File 1: {os.path.basename(file_path)}")
                self.col1_selector.clear()
                self.col1_selector.addItems(self.df1.columns.astype(str).tolist())
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file 1:\n{e}")

    def load_file2(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Second Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                self.df2 = pd.read_excel(file_path)
                self.file2_label.setText(f"File 2: {os.path.basename(file_path)}")
                self.col2_selector.clear()
                self.col2_selector.addItems(self.df2.columns.astype(str).tolist())
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file 2:\n{e}")

    def normalize_key(self, series):
        return (
            series.astype(str)
            .str.upper()
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
            .fillna("")
        )

    def compare_columns(self):
        if self.df1 is None or self.df2 is None:
            QMessageBox.warning(self, "Warning", "Please load both files first!")
            return

        col1 = self.col1_selector.currentText()
        col2 = self.col2_selector.currentText()
        if not col1 or not col2:
            QMessageBox.warning(self, "Warning", "Please select columns to compare!")
            return

        try:
            # Create normalized composite keys
            self.df1["Composite_Key"] = self.normalize_key(self.df1[col1])
            self.df2["Composite_Key"] = self.normalize_key(self.df2[col2])

            # Merge
            merged = pd.merge(
                self.df1, self.df2, on="Composite_Key",
                how="outer", suffixes=("_FILE1", "_FILE2"), indicator=True
            )

            only_in_file1 = merged[merged["_merge"] == "left_only"]
            only_in_file2 = merged[merged["_merge"] == "right_only"]
            common_values = merged[merged["_merge"] == "both"]

            # Store results
            self.results = {
                "only_in_file1": only_in_file1,
                "only_in_file2": only_in_file2,
                "common_values": common_values,
                "merged": merged
            }

            # Display results
            self.only_in_file1_text.setPlainText(only_in_file1.to_string(index=False))
            self.only_in_file2_text.setPlainText(only_in_file2.to_string(index=False))
            self.common_text.setPlainText(common_values.to_string(index=False))
            self.merged_text.setPlainText(merged.to_string(index=False))

            # Summary
            self.generate_summary(col1, col2, only_in_file1, only_in_file2, common_values)

            QMessageBox.information(self, "Success", "Comparison completed successfully!")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to compare columns:\n{e}")

    def generate_summary(self, col1, col2, only_in_df1, only_in_df2, common_values):
        total1 = len(only_in_df1) + len(common_values)
        total2 = len(only_in_df2) + len(common_values)

        summary_text = f"""
        <h2>Comparison Summary</h2>
        <p><b>Compared:</b> '{col1}' (File 1) vs '{col2}' (File 2)</p>
        <ul>
            <li><b>Only in first file:</b> {len(only_in_df1)}</li>
            <li><b>Only in second file:</b> {len(only_in_df2)}</li>
            <li><b>Common values:</b> {len(common_values)}</li>
            <li><b>Total unique in first file:</b> {total1}</li>
            <li><b>Total unique in second file:</b> {total2}</li>
        </ul>
        """
        self.summary_text.setHtml(summary_text)
        self.summary_text.moveCursor(QTextCursor.MoveOperation.Start)

    def export_results(self):
        if not self.results:
            QMessageBox.warning(self, "Warning", "No comparison results to export!")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    pd.DataFrame(self.results["only_in_file1"]).to_excel(writer, sheet_name="Only in First File", index=False)
                    pd.DataFrame(self.results["only_in_file2"]).to_excel(writer, sheet_name="Only in Second File", index=False)
                    pd.DataFrame(self.results["common_values"]).to_excel(writer, sheet_name="Common Values (Matched)", index=False)
                    pd.DataFrame(self.results["merged"]).to_excel(writer, sheet_name="Merged Results", index=False)

                    summary_df = pd.DataFrame({
                        "Comparison Summary": [
                            f"Only in first file: {len(self.results['only_in_file1'])}",
                            f"Only in second file: {len(self.results['only_in_file2'])}",
                            f"Common values: {len(self.results['common_values'])}"
                        ]
                    })
                    summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # Highlight differences in merged sheet
                wb = load_workbook(save_path)
                ws = wb["Merged Results"]
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row in ws.iter_rows(min_row=2):
                    if row[-1].value in ("left_only", "right_only"):
                        for cell in row:
                            cell.fill = fill
                wb.save(save_path)

                QMessageBox.information(self, "Success", f"Results exported successfully to:\n{save_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export results:\n{e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelComparator()
    window.show()
    sys.exit(app.exec())
