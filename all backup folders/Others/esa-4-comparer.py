import sys
import pandas as pd
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QFileDialog,
                             QComboBox, QTableWidget, QTableWidgetItem,
                             QMessageBox, QTabWidget, QHeaderView, QTextEdit,
                             QListWidget, QGroupBox, QScrollArea, QCheckBox,
                             QGridLayout, QSizePolicy)
from PyQt6.QtCore import Qt
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelComparator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.df1 = None
        self.df2 = None
        self.current_comparison_results = None

    def initUI(self):
        self.setWindowTitle('Excel Column Comparison Tool')
        self.setGeometry(100, 100, 1600, 1000)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # File selection section
        file_layout = QHBoxLayout()

        self.file1_btn = QPushButton('Select First Excel File')
        self.file1_btn.clicked.connect(self.load_file1)
        file_layout.addWidget(self.file1_btn)

        self.file1_label = QLabel('No file selected')
        file_layout.addWidget(self.file1_label)

        self.file2_btn = QPushButton('Select Second Excel File')
        self.file2_btn.clicked.connect(self.load_file2)
        file_layout.addWidget(self.file2_btn)

        self.file2_label = QLabel('No file selected')
        file_layout.addWidget(self.file2_label)

        layout.addLayout(file_layout)

        # Column selection section with scroll area
        scroll_widget = QWidget()
        scroll_layout = QHBoxLayout(scroll_widget)

        # First file columns to compare
        self.compare_group1 = QGroupBox("Select columns to compare from first file")
        self.compare_group1.setEnabled(False)
        compare_layout1 = QVBoxLayout(self.compare_group1)

        self.compare_list1 = QListWidget()
        self.compare_list1.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        compare_layout1.addWidget(self.compare_list1)

        # Second file columns to compare
        self.compare_group2 = QGroupBox("Select columns to compare from second file")
        self.compare_group2.setEnabled(False)
        compare_layout2 = QVBoxLayout(self.compare_group2)

        self.compare_list2 = QListWidget()
        self.compare_list2.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        compare_layout2.addWidget(self.compare_list2)

        scroll_layout.addWidget(self.compare_group1)
        scroll_layout.addWidget(self.compare_group2)

        # Additional columns to retain
        self.retain_group1 = QGroupBox("Additional columns to retain from first file")
        self.retain_group1.setEnabled(False)
        retain_layout1 = QVBoxLayout(self.retain_group1)

        self.retain_list1 = QListWidget()
        self.retain_list1.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        retain_layout1.addWidget(self.retain_list1)

        self.retain_group2 = QGroupBox("Additional columns to retain from second file")
        self.retain_group2.setEnabled(False)
        retain_layout2 = QVBoxLayout(self.retain_group2)

        self.retain_list2 = QListWidget()
        self.retain_list2.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        retain_layout2.addWidget(self.retain_list2)

        scroll_layout.addWidget(self.retain_group1)
        scroll_layout.addWidget(self.retain_group2)

        scroll_area = QScrollArea()
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Button section
        button_layout = QHBoxLayout()

        self.compare_btn = QPushButton('Compare Columns')
        self.compare_btn.clicked.connect(self.compare_columns)
        self.compare_btn.setEnabled(False)
        button_layout.addWidget(self.compare_btn)

        self.export_btn = QPushButton('Export to Excel')
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        button_layout.addWidget(self.export_btn)

        layout.addLayout(button_layout)

        # Tab widget for results
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Tab 1: Detailed comparison
        self.detail_tab = QWidget()
        self.detail_layout = QHBoxLayout(self.detail_tab)

        self.table1 = QTableWidget()
        self.table2 = QTableWidget()

        self.detail_layout.addWidget(self.table1)
        self.detail_layout.addWidget(self.table2)

        self.tabs.addTab(self.detail_tab, "Detailed Comparison")

        # Tab 2: Summary
        self.summary_tab = QWidget()
        self.summary_layout = QVBoxLayout(self.summary_tab)

        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_layout.addWidget(self.summary_text)

        self.tabs.addTab(self.summary_tab, "Summary")

        # Tab 3: Merged Results
        self.merged_tab = QWidget()
        self.merged_layout = QVBoxLayout(self.merged_tab)

        self.merged_table = QTableWidget()
        self.merged_layout.addWidget(self.merged_table)

        self.tabs.addTab(self.merged_tab, "Merged Results")

    def load_file1(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select First Excel File", "", "Excel Files (*.xlsx *.xls)")

        if file_path:
            try:
                self.df1 = pd.read_excel(file_path)
                self.file1_label.setText(file_path.split('/')[-1])
                self.populate_lists(self.compare_list1, self.df1)
                self.populate_lists(self.retain_list1, self.df1)
                self.compare_group1.setEnabled(True)
                self.retain_group1.setEnabled(True)
                self.check_comparison_readiness()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file: {str(e)}")

    def load_file2(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Second Excel File", "", "Excel Files (*.xlsx *.xls)")

        if file_path:
            try:
                self.df2 = pd.read_excel(file_path)
                self.file2_label.setText(file_path.split('/')[-1])
                self.populate_lists(self.compare_list2, self.df2)
                self.populate_lists(self.retain_list2, self.df2)
                self.compare_group2.setEnabled(True)
                self.retain_group2.setEnabled(True)
                self.check_comparison_readiness()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file: {str(e)}")

    def populate_lists(self, list_widget, df):
        list_widget.clear()
        if df is not None:
            list_widget.addItems(df.columns.tolist())

    def check_comparison_readiness(self):
        if self.df1 is not None and self.df2 is not None:
            self.compare_btn.setEnabled(True)
        else:
            self.compare_btn.setEnabled(False)

    def compare_columns(self):
        # Get selected columns to compare
        compare_cols1 = [item.text() for item in self.compare_list1.selectedItems()]
        compare_cols2 = [item.text() for item in self.compare_list2.selectedItems()]

        if not compare_cols1 or not compare_cols2:
            QMessageBox.warning(self, "Warning", "Please select at least one column to compare from each file")
            return

        try:
            # Get selected columns to retain
            retain_cols1 = [item.text() for item in self.retain_list1.selectedItems()]
            retain_cols2 = [item.text() for item in self.retain_list2.selectedItems()]

            # Create composite keys for comparison
            self.df1['Composite_Key'] = self.df1[compare_cols1].astype(str).agg('|'.join, axis=1)
            self.df2['Composite_Key'] = self.df2[compare_cols2].astype(str).agg('|'.join, axis=1)

            # Get unique values from both composite keys
            values1 = set(self.df1['Composite_Key'].dropna().unique())
            values2 = set(self.df2['Composite_Key'].dropna().unique())

            # Find differences
            only_in_df1 = values1 - values2
            only_in_df2 = values2 - values1
            common_values = values1 & values2

            # Prepare detailed data for export
            detailed_only_in_df1 = self.df1[self.df1['Composite_Key'].isin(only_in_df1)]
            detailed_only_in_df2 = self.df2[self.df2['Composite_Key'].isin(only_in_df2)]

            # For common values, we need to match rows between both files
            # Create a mapping from composite key to rows for efficient lookup
            df1_dict = {row['Composite_Key']: row for _, row in self.df1.iterrows()}
            df2_dict = {row['Composite_Key']: row for _, row in self.df2.iterrows()}

            # Create matched common data
            matched_common_data = []
            for key in common_values:
                if key in df1_dict and key in df2_dict:
                    matched_common_data.append({
                        'key': key,
                        'df1_row': df1_dict[key],
                        'df2_row': df2_dict[key]
                    })

            # Create merged results
            merged_results = self.create_merged_results(
                detailed_only_in_df1, detailed_only_in_df2,
                matched_common_data,
                compare_cols1, compare_cols2,
                retain_cols1, retain_cols2
            )

            # Store results for export
            self.current_comparison_results = {
                'compare_cols1': compare_cols1,
                'compare_cols2': compare_cols2,
                'only_in_df1': only_in_df1,
                'only_in_df2': only_in_df2,
                'common_values': common_values,
                'detailed_only_in_df1': detailed_only_in_df1,
                'detailed_only_in_df2': detailed_only_in_df2,
                'matched_common_data': matched_common_data,
                'retain_cols1': retain_cols1,
                'retain_cols2': retain_cols2,
                'merged_results': merged_results
            }

            # Display results
            self.display_results(self.table1, f"Values only in first file", only_in_df1)
            self.display_results(self.table2, f"Values only in second file", only_in_df2)

            # Display merged results
            self.display_merged_results(merged_results)

            # Generate summary
            self.generate_summary(compare_cols1, compare_cols2, only_in_df1, only_in_df2, common_values)

            # Enable export button
            self.export_btn.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Comparison failed: {str(e)}")

    def create_merged_results(self, only_in_df1, only_in_df2, matched_common_data,
                              compare_cols1, compare_cols2, retain_cols1, retain_cols2):
        # Create a merged dataframe with all data
        merged_df = pd.DataFrame()

        # Add rows that exist only in first file
        for _, row in only_in_df1.iterrows():
            new_row = {
                'Status': 'Only in File 1',
                'Composite_Key': row['Composite_Key']
            }

            # Add comparison columns from first file
            for col in compare_cols1:
                new_row[f'File1_{col}'] = row[col]

            # Add comparison columns from second file (empty)
            for col in compare_cols2:
                new_row[f'File2_{col}'] = ''

            # Add retained columns from first file
            for col in retain_cols1:
                if col in row:
                    new_row[f'File1_{col}'] = row[col]

            # Add retained columns from second file (empty)
            for col in retain_cols2:
                new_row[f'File2_{col}'] = ''

            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)

        # Add rows that exist only in second file
        for _, row in only_in_df2.iterrows():
            new_row = {
                'Status': 'Only in File 2',
                'Composite_Key': row['Composite_Key']
            }

            # Add comparison columns from first file (empty)
            for col in compare_cols1:
                new_row[f'File1_{col}'] = ''

            # Add comparison columns from second file
            for col in compare_cols2:
                new_row[f'File2_{col}'] = row[col]

            # Add retained columns from first file (empty)
            for col in retain_cols1:
                new_row[f'File1_{col}'] = ''

            # Add retained columns from second file
            for col in retain_cols2:
                if col in row:
                    new_row[f'File2_{col}'] = row[col]

            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)

        # Add rows that exist in both files (matched properly)
        for match in matched_common_data:
            row1 = match['df1_row']
            row2 = match['df2_row']

            new_row = {
                'Status': 'In Both Files',
                'Composite_Key': match['key']
            }

            # Add comparison columns from both files
            for col in compare_cols1:
                new_row[f'File1_{col}'] = row1[col]
            for col in compare_cols2:
                new_row[f'File2_{col}'] = row2[col]

            # Add retained columns from both files
            for col in retain_cols1:
                if col in row1:
                    new_row[f'File1_{col}'] = row1[col]
            for col in retain_cols2:
                if col in row2:
                    new_row[f'File2_{col}'] = row2[col]

            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)

        return merged_df

    def display_results(self, table, title, data):
        table.clear()
        table.setColumnCount(1)
        table.setRowCount(len(data))
        table.setHorizontalHeaderLabels([title])

        sorted_data = sorted(data)
        for row, value in enumerate(sorted_data):
            # Split the composite key back into individual values for display
            display_value = value.replace('|', ' | ')
            table.setItem(row, 0, QTableWidgetItem(display_value))

        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)

    def display_merged_results(self, merged_df):
        self.merged_table.clear()

        if merged_df.empty:
            self.merged_table.setRowCount(0)
            self.merged_table.setColumnCount(0)
            return

        # Set up table dimensions
        self.merged_table.setRowCount(len(merged_df))
        self.merged_table.setColumnCount(len(merged_df.columns))

        # Set headers
        self.merged_table.setHorizontalHeaderLabels(merged_df.columns)

        # Populate table
        for row_idx, (_, row) in enumerate(merged_df.iterrows()):
            for col_idx, col_name in enumerate(merged_df.columns):
                value = row[col_name]
                item = QTableWidgetItem(str(value) if pd.notna(value) else "")

                # Color code based on status
                if col_name == 'Status':
                    if value == 'Only in File 1':
                        item.setBackground(Qt.GlobalColor.yellow)
                    elif value == 'Only in File 2':
                        item.setBackground(Qt.GlobalColor.cyan)
                    elif value == 'In Both Files':
                        item.setBackground(Qt.GlobalColor.green)

                self.merged_table.setItem(row_idx, col_idx, item)

        # Resize columns to fit content
        self.merged_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    def generate_summary(self, compare_cols1, compare_cols2, only_in_df1, only_in_df2, common_values):
        summary_text = f"""
        <h2>Comparison Summary</h2>
        <p><b>Compared:</b> '{", ".join(compare_cols1)}' from first file vs '{", ".join(compare_cols2)}' from second file</p>
        <p><b>Values only in first file:</b> {len(only_in_df1)}</p>
        <p><b>Values only in second file:</b> {len(only_in_df2)}</p>
        <p><b>Common values:</b> {len(common_values)}</p>
        <p><b>Total unique combinations in first file:</b> {len(only_in_df1) + len(common_values)}</p>
        <p><b>Total unique combinations in second file:</b> {len(only_in_df2) + len(common_values)}</p>
        """

        self.summary_text.setHtml(summary_text)

    def export_to_excel(self):
        if not self.current_comparison_results:
            QMessageBox.warning(self, "Warning", "No comparison results to export")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Comparison Results", "", "Excel Files (*.xlsx)")

        if not file_path:
            return

        try:
            # Create a new workbook
            wb = openpyxl.Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Add summary information
            ws_summary['A1'] = "Comparison Summary"
            ws_summary['A1'].font = Font(bold=True, size=14)

            ws_summary[
                'A3'] = f"Compared: '{', '.join(self.current_comparison_results['compare_cols1'])}' from first file vs '{', '.join(self.current_comparison_results['compare_cols2'])}' from second file"
            ws_summary['A5'] = f"Values only in first file: {len(self.current_comparison_results['only_in_df1'])}"
            ws_summary['A6'] = f"Values only in second file: {len(self.current_comparison_results['only_in_df2'])}"
            ws_summary['A7'] = f"Common values: {len(self.current_comparison_results['common_values'])}"
            ws_summary[
                'A8'] = f"Total unique combinations in first file: {len(self.current_comparison_results['only_in_df1']) + len(self.current_comparison_results['common_values'])}"
            ws_summary[
                'A9'] = f"Total unique combinations in second file: {len(self.current_comparison_results['only_in_df2']) + len(self.current_comparison_results['common_values'])}"

            # Only in first file sheet (with additional columns)
            ws_only1 = wb.create_sheet("Only in First File")
            self.create_detailed_sheet(
                ws_only1,
                f"Values only in first file",
                self.current_comparison_results['detailed_only_in_df1'],
                self.current_comparison_results['compare_cols1'],
                self.current_comparison_results['retain_cols1']
            )

            # Only in second file sheet (with additional columns)
            ws_only2 = wb.create_sheet("Only in Second File")
            self.create_detailed_sheet(
                ws_only2,
                f"Values only in second file",
                self.current_comparison_results['detailed_only_in_df2'],
                self.current_comparison_results['compare_cols2'],
                self.current_comparison_results['retain_cols2']
            )

            # Common values sheet (matched data)
            ws_common = wb.create_sheet("Common Values (Matched)")
            self.create_matched_common_sheet(
                ws_common,
                self.current_comparison_results['matched_common_data'],
                self.current_comparison_results['compare_cols1'],
                self.current_comparison_results['compare_cols2'],
                self.current_comparison_results['retain_cols1'],
                self.current_comparison_results['retain_cols2']
            )

            # Merged results sheet
            ws_merged = wb.create_sheet("Merged Results")
            self.create_merged_sheet(
                ws_merged,
                self.current_comparison_results['merged_results']
            )

            # Save the workbook
            wb.save(file_path)

            QMessageBox.information(self, "Success", f"Results exported to {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export results: {str(e)}")

    def create_detailed_sheet(self, worksheet, title, data, compare_columns, retain_columns):
        # Add title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=12)

        # Determine which columns to include
        all_columns = compare_columns + retain_columns
        columns_to_include = [col for col in all_columns if col in data.columns]

        # Write headers
        for col_idx, column in enumerate(columns_to_include, start=1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.value = column
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        for row_idx, (_, row) in enumerate(data.iterrows(), start=4):
            for col_idx, column in enumerate(columns_to_include, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = row[column] if pd.notna(row[column]) else ""

        # Auto-adjust column widths
        for col_idx, column in enumerate(columns_to_include, start=1):
            max_length = max(
                len(str(column)),  # Header length
                *[len(str(row[column])) if pd.notna(row[column]) else 0 for _, row in data.iterrows()]  # Data lengths
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

    def create_matched_common_sheet(self, worksheet, matched_data, compare_cols1, compare_cols2, retain_cols1,
                                    retain_cols2):
        if not matched_data:
            return

        # Add title
        worksheet['A1'] = "Common Values (Matched Rows)"
        worksheet['A1'].font = Font(bold=True, size=12)

        # Create headers
        headers = ['Composite_Key', 'Status']

        # Add comparison columns from both files
        for col in compare_cols1:
            headers.append(f'File1_{col}')
        for col in compare_cols2:
            headers.append(f'File2_{col}')

        # Add retained columns from both files
        for col in retain_cols1:
            headers.append(f'File1_{col}')
        for col in retain_cols2:
            headers.append(f'File2_{col}')

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        for row_idx, match in enumerate(matched_data, start=4):
            row1 = match['df1_row']
            row2 = match['df2_row']

            # Composite Key
            worksheet.cell(row=row_idx, column=1).value = match['key']

            # Status
            worksheet.cell(row=row_idx, column=2).value = 'In Both Files'

            # Comparison columns from first file
            col_idx = 3
            for col in compare_cols1:
                worksheet.cell(row=row_idx, column=col_idx).value = row1[col] if pd.notna(row1[col]) else ""
                col_idx += 1

            # Comparison columns from second file
            for col in compare_cols2:
                worksheet.cell(row=row_idx, column=col_idx).value = row2[col] if pd.notna(row2[col]) else ""
                col_idx += 1

            # Retained columns from first file
            for col in retain_cols1:
                if col in row1:
                    worksheet.cell(row=row_idx, column=col_idx).value = row1[col] if pd.notna(row1[col]) else ""
                col_idx += 1

            # Retained columns from second file
            for col in retain_cols2:
                if col in row2:
                    worksheet.cell(row=row_idx, column=col_idx).value = row2[col] if pd.notna(row2[col]) else ""
                col_idx += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = max(
                len(str(headers[col_idx - 1])),  # Header length
                *[len(str(worksheet.cell(row=row_idx, column=col_idx).value or ""))
                  for row_idx in range(4, len(matched_data) + 4)]  # Data lengths
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

    def create_merged_sheet(self, worksheet, merged_df):
        if merged_df.empty:
            return

        # Write headers
        for col_idx, column in enumerate(merged_df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column
            cell.font = Font(bold=True)

            # Color code status column header
            if column == 'Status':
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        for row_idx, (_, row) in enumerate(merged_df.iterrows(), start=2):
            for col_idx, column in enumerate(merged_df.columns, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = row[column] if pd.notna(row[column]) else ""

                # Color code based on status
                if column == 'Status':
                    if row[column] == 'Only in File 1':
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                    elif row[column] == 'Only in File 2':
                        cell.fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Cyan
                    elif row[column] == 'In Both Files':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

        # Auto-adjust column widths
        for col_idx, column in enumerate(merged_df.columns, start=1):
            max_length = max(
                len(str(column)),  # Header length
                *[len(str(row[column])) if pd.notna(row[column]) else 0 for _, row in merged_df.iterrows()]
                # Data lengths
            )
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)


def main():
    app = QApplication(sys.argv)
    window = ExcelComparator()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()