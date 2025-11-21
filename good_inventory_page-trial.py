import pandas as pd
import os
import sys
import traceback
from datetime import datetime
import qtawesome as fa

# --- Email Imports (dotenv is no longer needed) ---
from email.message import EmailMessage
from io import BytesIO
import smtplib
import ssl

# --- PyQt6 Imports ---
# <<< MODIFIED: Added QSettings and new Dialog-related widgets >>>
from PyQt6.QtCore import Qt, QObject, pyqtSignal, QThread, QDate, QSize, QSettings
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QTableWidget,
                             QTableWidgetItem, QHeaderView, QMessageBox, QHBoxLayout,
                             QLabel, QPushButton, QLineEdit, QAbstractItemView,
                             QDateEdit, QGroupBox, QFileDialog, QTabWidget,
                             QGridLayout, QProgressBar, QComboBox, QDialog,
                             QFormLayout, QDialogButtonBox, QSpinBox)

# --- SQLAlchemy and OpenPyXL Imports (unchanged) ---
from sqlalchemy import create_engine, text, Engine
from sqlalchemy.engine import URL
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- UI CONSTANTS (unchanged) ---
PRIMARY_ACCENT_COLOR = "#007bff"
PRIMARY_ACCENT_HOVER = "#e9f0ff"
NEUTRAL_COLOR = "#6c757d"
LIGHT_TEXT_COLOR = "#333333"
BACKGROUND_CONTENT_COLOR = "#f4f7fc"
INPUT_BACKGROUND_COLOR = "#ffffff"
GROUP_BOX_HEADER_COLOR = "#f4f7fc"
TABLE_SELECTION_COLOR = "#3a506b"
ERROR_COLOR = "#dc3545"
HEADER_AND_ICON_COLOR = "#3a506b"
TABLE_HEADER_TEXT_COLOR = "#4f4f4f"


# --- Custom UpperCaseLineEdit Widget (unchanged) ---
class UpperCaseLineEdit(QLineEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.textEdited.connect(self._to_upper)

    def _to_upper(self, text: str):
        if text and text != self.text().upper():
            self.blockSignals(True);
            self.setText(text.upper());
            self.blockSignals(False)


# --- Worker for Database Calculations (unchanged) ---
class InventoryWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, product_filter: str, lot_filter: str, as_of_date: str, fg_type_filter: str):
        super().__init__()
        self.engine = engine;
        self.product_filter = product_filter;
        self.lot_filter = lot_filter
        self.as_of_date = as_of_date;
        self.fg_type_filter = fg_type_filter

    def run(self):
        try:
            params = {'as_of_date': self.as_of_date}
            date_filter_clause = "AND transaction_date <= :as_of_date"
            product_filter_clause = ""
            lot_filter_clause = ""
            fg_type_clause = ""
            if self.product_filter:
                product_filter_clause = "AND product_code = :product";
                params['product'] = self.product_filter
            if self.lot_filter:
                lot_filter_clause = "AND lot_number = :lot";
                params['lot'] = self.lot_filter
            if self.fg_type_filter in ["MB", "DC"]:
                fg_type_clause = "AND fg_type = :fg_type";
                params['fg_type'] = self.fg_type_filter
            query_str = f"""
                WITH combined_movements AS (
                    SELECT UPPER(TRIM(product_code)) AS product_code, UPPER(TRIM(lot_number)) AS lot_number, COALESCE(UPPER(TRIM(fg_type)), CASE WHEN product_code LIKE '%-%' THEN 'DC' ELSE 'MB' END) AS fg_type, COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in, 0.0 AS quantity_out FROM beginv_sheet1 WHERE product_code IS NOT NULL AND TRIM(product_code) <> '' AND lot_number IS NOT NULL AND TRIM(lot_number) <> ''
                    UNION ALL
                    SELECT UPPER(TRIM(product_code)) AS product_code, UPPER(TRIM(lot_number)) AS lot_number, CASE WHEN product_code LIKE '%-%' THEN 'DC' ELSE 'MB' END AS fg_type, COALESCE(CAST(quantity_in AS NUMERIC), 0) AS quantity_in, COALESCE(CAST(quantity_out AS NUMERIC), 0) AS quantity_out FROM transactions WHERE product_code IS NOT NULL AND TRIM(product_code) <> '' AND lot_number IS NOT NULL AND TRIM(lot_number) <> '' {date_filter_clause}
                ), lot_details AS (
                    SELECT UPPER(TRIM(lot_number)) as lot_number, MAX(UPPER(TRIM(location))) as location, MAX(COALESCE(UPPER(TRIM(bag_number)), UPPER(TRIM(box_number)), '')) as bag_box_number FROM beginv_sheet1 WHERE location IS NOT NULL AND TRIM(location) <> '' GROUP BY UPPER(TRIM(lot_number))
                ), lot_summary AS (
                    SELECT c.lot_number, MAX(c.product_code) AS product_code, MAX(c.fg_type) AS fg_type, COALESCE(SUM(c.quantity_in), 0) - COALESCE(SUM(c.quantity_out), 0) AS current_balance, MAX(det.location) as location, MAX(det.bag_box_number) as bag_box_number FROM combined_movements c LEFT JOIN lot_details det ON c.lot_number = det.lot_number GROUP BY c.lot_number HAVING (COALESCE(SUM(c.quantity_in), 0) - COALESCE(SUM(c.quantity_out), 0)) > 0.001
                )
                SELECT product_code, lot_number, current_balance, location, bag_box_number, fg_type FROM lot_summary WHERE location IS NOT NULL {product_filter_clause} {lot_filter_clause} {fg_type_clause} ORDER BY product_code, lot_number;
            """
            with self.engine.connect() as conn:
                results = conn.execute(text(query_str), params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame(
                columns=['product_code', 'lot_number', 'current_balance', 'location', 'bag_box_number', 'fg_type'])
            if not df.empty:
                df['current_balance'] = pd.to_numeric(df['current_balance'])
            self.finished.emit(df)
        except Exception:
            self.error.emit("Database query failed.", traceback.format_exc())


# --- Reusable Error Dialog (unchanged) ---
def show_error_message(parent, title, message, detailed_text=""):
    msg_box = QMessageBox(parent)
    msg_box.setIcon(QMessageBox.Icon.Critical)
    msg_box.setWindowTitle(title)
    msg_box.setText(f"<b>{message}</b>")
    if detailed_text:
        msg_box.setDetailedText(detailed_text)
    msg_box.exec()


# <<< NEW: SETTINGS DIALOG CLASS >>>
class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Email Settings")
        self.setMinimumWidth(450)

        # Use QSettings to store data persistently by company and app name
        self.settings = QSettings("MyCompany", "FGInventoryApp")

        main_layout = QVBoxLayout(self)

        sender_group = QGroupBox("Sender Details (SMTP)")
        sender_layout = QFormLayout(sender_group)
        self.sender_email_input = QLineEdit()
        self.sender_password_input = QLineEdit()
        self.sender_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.smtp_server_input = QLineEdit()
        self.smtp_port_input = QSpinBox()
        self.smtp_port_input.setRange(1, 65535)
        sender_layout.addRow("Sender Email:", self.sender_email_input)
        sender_layout.addRow("Password:", self.sender_password_input)
        sender_layout.addRow("SMTP Server:", self.smtp_server_input)
        sender_layout.addRow("SMTP Port:", self.smtp_port_input)
        main_layout.addWidget(sender_group)

        recipient_group = QGroupBox("Default Recipient")
        recipient_layout = QFormLayout(recipient_group)
        self.recipient_email_input = QLineEdit()
        self.recipient_email_input.setPlaceholderText("e.g., user1@example.com, user2@example.com")
        recipient_layout.addRow("Recipient Email(s):", self.recipient_email_input)
        main_layout.addWidget(recipient_group)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        main_layout.addWidget(button_box)

        self.load_settings()

    def load_settings(self):
        """Load settings from QSettings and populate the fields."""
        self.sender_email_input.setText(self.settings.value("email/sender_email", ""))
        if self.settings.value("email/sender_password"):
            self.sender_password_input.setPlaceholderText("Password is saved. Enter new to change.")
        self.smtp_server_input.setText(self.settings.value("email/smtp_server", ""))
        self.smtp_port_input.setValue(int(self.settings.value("email/smtp_port", 587)))
        self.recipient_email_input.setText(self.settings.value("email/recipient_email", ""))

    def accept(self):
        """Save the current settings and close the dialog."""
        self.settings.setValue("email/sender_email", self.sender_email_input.text().strip())
        if self.sender_password_input.text():
            self.settings.setValue("email/sender_password", self.sender_password_input.text())
        self.settings.setValue("email/smtp_server", self.smtp_server_input.text().strip())
        self.settings.setValue("email/smtp_port", self.smtp_port_input.value())
        self.settings.setValue("email/recipient_email", self.recipient_email_input.text().strip())
        QMessageBox.information(self, "Success", "Settings have been saved.")
        super().accept()


# <<< MODIFIED: EMAIL FUNCTION NOW USES A CONFIG DICTIONARY >>>
def send_email_with_excel(df_dict, email_config):
    """Creates and sends an email using a dictionary of settings."""
    sender_email = email_config.get('sender_email')
    sender_password = email_config.get('sender_password')
    smtp_server = email_config.get('smtp_server')
    smtp_port = email_config.get('smtp_port')
    recipient_email = email_config.get('recipient_email')
    subject = email_config.get('subject')
    excel_fn = email_config.get('excel_fn')

    if not all([sender_email, sender_password, smtp_server, smtp_port, recipient_email]):
        raise ValueError("Incomplete email settings. Please configure them in the Settings menu.")

    recipients = [email.strip() for email in recipient_email.split(',')]

    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    excel_buffer.seek(0)

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipients)
    msg.set_content(f'Please find the attached Excel report: {excel_fn}.\n\nThis is an automated message.')

    msg.add_attachment(
        excel_buffer.read(),
        maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=excel_fn)

    context = ssl.create_default_context()
    if str(smtp_port) == '465':
        with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg, from_addr=sender_email, to_addrs=recipients)
    else:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls(context=context)
            server.login(sender_email, sender_password)
            server.send_message(msg, from_addr=sender_email, to_addrs=recipients)


# --- Advanced Dashboard Widget (unchanged) ---
class DashboardWidget(QWidget):
    # ... (Omitted for brevity, no changes needed)
    def __init__(self):
        super().__init__()
        self.data_df = pd.DataFrame()
        self.init_ui()
        self.setStyleSheet(self._get_dashboard_styles())

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(15)
        grid_layout = QGridLayout()
        grid_layout.setRowStretch(2, 1)
        grid_layout.setColumnStretch(0, 1)
        grid_layout.setColumnStretch(1, 1)
        summary_group = QGroupBox("Overall Summary Metrics")
        summary_group.setObjectName("SummaryGroup")
        summary_layout = QHBoxLayout(summary_group)
        self.total_lots_label = self._create_summary_box("Total Lots:", "0")
        self.total_products_label = self._create_summary_box("Unique Products:", "0")
        self.overall_balance_label = self._create_summary_box("Overall Balance (kg):", "0.00")
        summary_layout.addWidget(self.total_lots_label)
        summary_layout.addWidget(self.total_products_label)
        summary_layout.addWidget(self.overall_balance_label)
        grid_layout.addWidget(summary_group, 0, 0, 1, 2)
        self.contribution_table = self._create_contribution_table("product")
        contribution_group = QGroupBox("Top 10 Product Contribution (by Mass)")
        contribution_group.setObjectName("ContributionGroup")
        vbox_contribution = QVBoxLayout(contribution_group)
        vbox_contribution.addWidget(self.contribution_table)
        grid_layout.addWidget(contribution_group, 1, 0)
        self.location_table = self._create_contribution_table("location")
        location_group = QGroupBox("Top 5 Locations by Mass")
        location_group.setObjectName("LocationGroup")
        vbox_location = QVBoxLayout(location_group)
        vbox_location.addWidget(self.location_table)
        grid_layout.addWidget(location_group, 1, 1)
        self.lot_stats_group = self._create_lot_statistics_group()
        grid_layout.addWidget(self.lot_stats_group, 2, 0)
        main_layout.addLayout(grid_layout)

    def _get_dashboard_styles(self) -> str:
        return f"""QGroupBox#SummaryGroup, QGroupBox#ContributionGroup, QGroupBox#LotStatsGroup, QGroupBox#LocationGroup{{border: 1px solid #e0e5eb; border-radius: 8px;margin-top: 12px; background-color: {INPUT_BACKGROUND_COLOR};}} QGroupBox::title{{subcontrol-origin: margin; subcontrol-position: top left;padding: 2px 10px; background-color: {GROUP_BOX_HEADER_COLOR};border: 1px solid #e0e5eb; border-bottom: none;border-top-left-radius: 8px; border-top-right-radius: 8px;font-weight: bold; color: {HEADER_AND_ICON_COLOR};}} QLabel#TitleLabel{{ font-size: 10pt; color: {HEADER_AND_ICON_COLOR}; background-color: transparent; }} QLabel#ValueLabel{{ font-size: 16pt; font-weight: bold; color: {HEADER_AND_ICON_COLOR}; background-color: transparent; }} QTableWidget{{ border: 1px solid #e0e5eb; background-color: {INPUT_BACKGROUND_COLOR}; gridline-color: #f0f3f8;}} QTableWidget::item:selected{{ background-color: {TABLE_SELECTION_COLOR}; color: white;}} QHeaderView::section{{ background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR};}} QProgressBar::chunk{{ background-color: {PRIMARY_ACCENT_COLOR}; border-radius: 4px; }}"""

    def _create_summary_box(self, title: str, initial_value: str) -> QWidget:
        widget = QWidget();
        layout = QVBoxLayout(widget);
        layout.setContentsMargins(15, 10, 15, 10)
        title_label = QLabel(title, objectName="TitleLabel");
        value_label = QLabel(initial_value, objectName="ValueLabel", alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label, alignment=Qt.AlignmentFlag.AlignCenter);
        layout.addWidget(value_label, alignment=Qt.AlignmentFlag.AlignCenter)
        widget.setStyleSheet(
            f"background-color: {INPUT_BACKGROUND_COLOR}; border: 1px solid #d1d9e6; border-radius: 8px;");
        return widget

    def _create_contribution_table(self, table_type: str) -> QTableWidget:
        table = QTableWidget(columnCount=4);
        header_label = "Product" if table_type == "product" else "Location"
        table.setHorizontalHeaderLabels([header_label, "Balance (kg)", "% Share", ""]);
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows);
        table.verticalHeader().setVisible(False);
        table.setShowGrid(True);
        table.setAlternatingRowColors(True)
        header = table.horizontalHeader();
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch);
        return table

    def _create_lot_statistics_group(self) -> QGroupBox:
        group = QGroupBox("Lot Size Statistics", objectName="LotStatsGroup");
        layout = QGridLayout(group)
        self.lot_stats = {'max_lot': QLabel("N/A"), 'min_lot': QLabel("N/A"), 'avg_lot': QLabel("N/A"),
                          'median_lot': QLabel("N/A")}
        label_style = f"font-size:10pt;color:{HEADER_AND_ICON_COLOR};background-color:transparent;";
        value_style = f"font-weight:bold;color:{HEADER_AND_ICON_COLOR};background-color:transparent;"
        layout.addWidget(QLabel("Largest Lot (kg):", styleSheet=label_style), 0, 0);
        self.lot_stats['max_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['max_lot'], 0, 1, alignment=Qt.AlignmentFlag.AlignRight)
        layout.addWidget(QLabel("Smallest Lot (kg):", styleSheet=label_style), 1, 0);
        self.lot_stats['min_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['min_lot'], 1, 1, alignment=Qt.AlignmentFlag.AlignRight)
        layout.addWidget(QLabel("Average Lot (kg):", styleSheet=label_style), 2, 0);
        self.lot_stats['avg_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['avg_lot'], 2, 1, alignment=Qt.AlignmentFlag.AlignRight)
        layout.addWidget(QLabel("Median Lot (kg):", styleSheet=label_style), 3, 0);
        self.lot_stats['median_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['median_lot'], 3, 1, alignment=Qt.AlignmentFlag.AlignRight)
        layout.setRowStretch(4, 1);
        return group

    def update_dashboard(self, df: pd.DataFrame):
        self.data_df = df;
        is_empty = df.empty
        self.total_lots_label.findChild(QLabel, "ValueLabel").setText("0" if is_empty else f"{len(df):,}")
        self.total_products_label.findChild(QLabel, "ValueLabel").setText(
            "0" if is_empty else f"{df['product_code'].nunique():,}")
        overall_balance = df['current_balance'].sum() if not is_empty else 0.0;
        self.overall_balance_label.findChild(QLabel, "ValueLabel").setText(f"{overall_balance:,.2f}")
        self.lot_stats['max_lot'].setText("N/A" if is_empty else f"{df['current_balance'].max():,.2f}");
        self.lot_stats['min_lot'].setText("N/A" if is_empty else f"{df['current_balance'].min():,.2f}")
        self.lot_stats['avg_lot'].setText("N/A" if is_empty else f"{df['current_balance'].mean():,.2f}");
        self.lot_stats['median_lot'].setText("N/A" if is_empty else f"{df['current_balance'].median():,.2f}")
        self.contribution_table.setRowCount(0);
        self.location_table.setRowCount(0)
        if is_empty: return
        self._populate_summary_table(df, self.contribution_table, 'product_code', overall_balance, 10)
        self._populate_summary_table(df, self.location_table, 'location', overall_balance, 5)

    def _populate_summary_table(self, df, table_widget, group_by_col, total_val, top_n):
        summary = df.groupby(group_by_col)['current_balance'].sum().nlargest(top_n).reset_index()
        summary['percentage'] = (summary['current_balance'] / total_val) * 100 if total_val > 0 else 0
        table_widget.setRowCount(len(summary))
        for i, row in summary.iterrows():
            table_widget.setItem(i, 0, QTableWidgetItem(str(row[group_by_col])));
            balance_item = QTableWidgetItem(f"{row['current_balance']:,.2f}");
            balance_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            table_widget.setItem(i, 1, balance_item);
            percent_item = QTableWidgetItem(f"{row['percentage']:.1f}%");
            percent_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            table_widget.setItem(i, 2, percent_item);
            progress_bar = QProgressBar(maximum=100, value=int(row['percentage']), textVisible=False);
            table_widget.setCellWidget(i, 3, progress_bar)


# --- Main Application Page ---
class GoodInventoryPage(QWidget):
    """The main widget containing all UI elements for the inventory tool."""

    def __init__(self, engine: Engine):
        super().__init__()
        self.engine = engine
        self.inventory_thread: QThread | None = None
        self.inventory_worker: InventoryWorker | None = None
        self.current_inventory_df = pd.DataFrame()
        self.init_ui()
        self.setStyleSheet(self._get_styles())
        self._start_inventory_calculation()

    def _get_styles(self) -> str:
        return f"""QWidget{{background-color:{BACKGROUND_CONTENT_COLOR}; color:{LIGHT_TEXT_COLOR}; font-family:'Segoe UI',Arial,sans-serif;}} QGroupBox{{border:1px solid #e0e5eb; border-radius:8px; margin-top:12px; background-color:{INPUT_BACKGROUND_COLOR};}} QGroupBox::title{{subcontrol-origin:margin; subcontrol-position:top left; padding:2px 10px; background-color:{GROUP_BOX_HEADER_COLOR}; border:1px solid #e0e5eb; border-bottom:1px solid {INPUT_BACKGROUND_COLOR}; border-top-left-radius:8px; border-top-right-radius:8px; font-weight:bold; color:#4f4f4f;}} QGroupBox QLabel{{background-color: transparent;}} QLabel#PageHeader{{font-size:15pt; font-weight:bold; color:{HEADER_AND_ICON_COLOR}; background-color:transparent;}} QLineEdit, QDateEdit, QComboBox{{border:1px solid #d1d9e6; padding:8px; border-radius:5px; background-color:{INPUT_BACKGROUND_COLOR};}} QLineEdit:focus, QDateEdit:focus, QComboBox:focus{{border:1px solid {PRIMARY_ACCENT_COLOR};}} QPushButton{{border:1px solid #d1d9e6; padding:8px 15px; border-radius:6px; font-weight:bold; background-color:{INPUT_BACKGROUND_COLOR};}} QPushButton:hover{{background-color:#f0f3f8;}} QPushButton#PrimaryButton{{color:{HEADER_AND_ICON_COLOR}; border:1px solid {HEADER_AND_ICON_COLOR};}} QPushButton#PrimaryButton:hover{{background-color:#e9f0ff;}} QTabWidget::pane{{border:1px solid #e0e5eb; border-radius:8px; background-color:{INPUT_BACKGROUND_COLOR}; padding:10px; margin-top:-1px;}} QTabBar::tab{{background:#e9eff7; color:{NEUTRAL_COLOR}; padding:8px 15px; border:1px solid #e0e5eb; border-bottom:none; border-top-left-radius:6px; border-top-right-radius:6px;}} QTabBar::tab:selected{{color:{HEADER_AND_ICON_COLOR}; background:{INPUT_BACKGROUND_COLOR}; border-bottom-color:{INPUT_BACKGROUND_COLOR}; font-weight:bold;}} QTableWidget{{border:1px solid #e0e5eb; background-color:{INPUT_BACKGROUND_COLOR}; selection-behavior:SelectRows; gridline-color: #f0f3f8;}} QTableWidget::item:hover{{background-color: transparent;}} QTableWidget::item:selected{{background-color:{TABLE_SELECTION_COLOR}; color:white;}} QHeaderView::section{{background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR};}} QLabel#TotalBalanceLabel{{background-color:{INPUT_BACKGROUND_COLOR}; color:{PRIMARY_ACCENT_COLOR}; padding:10px; border-radius:6px; border:1px solid #e0e5eb; font-weight:bold;}}"""

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        icon_label = QLabel()
        icon_label.setPixmap(fa.icon('fa5s.boxes', color=HEADER_AND_ICON_COLOR).pixmap(QSize(28, 28)))
        header_layout.addWidget(icon_label)
        header_layout.addWidget(QLabel("FG Inventory Computation and Export", objectName="PageHeader"))
        header_layout.addStretch()
        main_layout.addWidget(header_widget)
        main_layout.addWidget(self._create_filter_controls())
        self.tab_widget = QTabWidget()
        self.tab_widget.setIconSize(QSize(16, 16))
        self.inventory_tab = QWidget()
        inventory_layout = QVBoxLayout(self.inventory_tab)
        self.inventory_table = self._create_inventory_table()
        self.total_balance_label = QLabel("Total Balance: 0.00 kg", objectName="TotalBalanceLabel",
                                          alignment=Qt.AlignmentFlag.AlignRight)
        inventory_layout.addWidget(self.inventory_table)
        inventory_layout.addWidget(self.total_balance_label)
        self.tab_widget.addTab(self.inventory_tab, fa.icon('fa5.list-alt', color=HEADER_AND_ICON_COLOR),
                               "Inventory Details")
        self.dashboard_widget = DashboardWidget()
        self.tab_widget.addTab(self.dashboard_widget, fa.icon('fa5s.chart-area', color=HEADER_AND_ICON_COLOR),
                               "Dashboard")
        main_layout.addWidget(self.tab_widget, 1)

    def _create_filter_controls(self) -> QGroupBox:
        group = QGroupBox("Filters & Actions")
        layout = QHBoxLayout(group)
        layout.addWidget(QLabel("As Of Date:"))
        self.date_picker = QDateEdit(calendarPopup=True, date=QDate.currentDate(), displayFormat="yyyy-MM-dd")
        layout.addWidget(self.date_picker)
        layout.addSpacing(10)
        layout.addWidget(QLabel("FG Type:"))
        self.fg_type_combo = QComboBox();
        self.fg_type_combo.addItems(["All", "MB", "DC"])
        layout.addWidget(self.fg_type_combo)
        layout.addSpacing(10)
        layout.addWidget(QLabel("Product:"))
        self.product_code_input = UpperCaseLineEdit(placeholderText="Filter by Product Code")
        layout.addWidget(self.product_code_input, 1)
        layout.addSpacing(10)
        layout.addWidget(QLabel("Lot:"))
        self.lot_number_input = UpperCaseLineEdit(placeholderText="Filter by Lot Number")
        layout.addWidget(self.lot_number_input, 1)

        # <<< MODIFIED: Action buttons area >>>
        layout.addStretch()  # Pushes buttons to the right
        self.refresh_button = QPushButton("Refresh", objectName="PrimaryButton",
                                          icon=fa.icon('fa5s.sync-alt', color=HEADER_AND_ICON_COLOR))
        self.export_button = QPushButton("Export to Excel", objectName="PrimaryButton",
                                         icon=fa.icon('fa5s.file-excel', color=HEADER_AND_ICON_COLOR))
        self.email_button = QPushButton("Export & Email", objectName="PrimaryButton",
                                        icon=fa.icon('fa5s.paper-plane', color=HEADER_AND_ICON_COLOR))
        self.settings_button = QPushButton("Settings", objectName="PrimaryButton",
                                           icon=fa.icon('fa5s.cog', color=HEADER_AND_ICON_COLOR))

        layout.addWidget(self.refresh_button)
        layout.addWidget(self.export_button)
        layout.addWidget(self.email_button)
        layout.addWidget(self.settings_button)

        self.refresh_button.clicked.connect(self._start_inventory_calculation)
        self.product_code_input.returnPressed.connect(self._start_inventory_calculation)
        self.lot_number_input.returnPressed.connect(self._start_inventory_calculation)
        self.date_picker.dateChanged.connect(self._start_inventory_calculation)
        self.fg_type_combo.currentIndexChanged.connect(self._start_inventory_calculation)
        self.export_button.clicked.connect(self._export_to_excel)
        self.email_button.clicked.connect(self._export_and_email)
        self.settings_button.clicked.connect(self._open_settings_dialog)  # <<< CONNECTED
        return group

    def _create_inventory_table(self) -> QTableWidget:
        # ... (Omitted for brevity, unchanged)
        table = QTableWidget(columnCount=5);
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers);
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.verticalHeader().setVisible(False);
        table.setAlternatingRowColors(True);
        table.setHorizontalHeaderLabels(["PRODUCT", "LOT NUMBER", "QTY (kg)", "BAG/BOX NO.", "LOCATION"])
        header = table.horizontalHeader();
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents);
        return table

    def _start_inventory_calculation(self):
        # ... (Omitted for brevity, unchanged)
        if self.inventory_thread and self.inventory_thread.isRunning(): return
        self.set_controls_enabled(False);
        self._show_loading_state(self.inventory_table, "Calculating inventory balance...");
        self.dashboard_widget.update_dashboard(pd.DataFrame())
        self.inventory_thread = QThread();
        self.inventory_worker = InventoryWorker(engine=self.engine,
                                                product_filter=self.product_code_input.text().strip().upper(),
                                                lot_filter=self.lot_number_input.text().strip().upper(),
                                                as_of_date=self.date_picker.date().toString(Qt.DateFormat.ISODate),
                                                fg_type_filter=self.fg_type_combo.currentText())
        self.inventory_worker.moveToThread(self.inventory_thread);
        self.inventory_thread.started.connect(self.inventory_worker.run);
        self.inventory_worker.finished.connect(self._on_inventory_finished)
        self.inventory_worker.error.connect(self._on_calculation_error);
        self.inventory_worker.finished.connect(self.inventory_thread.quit);
        self.inventory_worker.error.connect(self.inventory_thread.quit)
        self.inventory_thread.finished.connect(self.inventory_worker.deleteLater);
        self.inventory_thread.finished.connect(self.inventory_thread.deleteLater);
        self.inventory_thread.finished.connect(self._reset_thread_state);
        self.inventory_thread.start()

    def _reset_thread_state(self):
        self.inventory_thread = None;
        self.inventory_worker = None

    def set_controls_enabled(self, enabled: bool):
        # ... (Omitted for brevity, unchanged)
        for widget in [self.lot_number_input, self.product_code_input, self.refresh_button, self.date_picker,
                       self.fg_type_combo]: widget.setEnabled(enabled)
        is_data_present = not self.current_inventory_df.empty;
        self.export_button.setEnabled(enabled and is_data_present);
        self.email_button.setEnabled(enabled and is_data_present)

    def _on_inventory_finished(self, df: pd.DataFrame):
        # ... (Omitted for brevity, unchanged)
        self.current_inventory_df = df.copy();
        self._display_inventory(df);
        self.dashboard_widget.update_dashboard(df);
        self.set_controls_enabled(True)

    def _on_calculation_error(self, error_message: str, detailed_traceback: str):
        # ... (Omitted for brevity, unchanged)
        show_error_message(self, "Calculation Error", error_message, detailed_traceback);
        self._show_loading_state(self.inventory_table, "Error during calculation.")
        self.dashboard_widget.update_dashboard(pd.DataFrame());
        self.set_controls_enabled(True)

    def _show_loading_state(self, table: QTableWidget, message: str):
        # ... (Omitted for brevity, unchanged)
        table.setRowCount(1);
        table.setSpan(0, 0, 1, table.columnCount());
        loading_item = QTableWidgetItem(message);
        loading_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter);
        table.setItem(0, 0, loading_item)

    def _display_inventory(self, df: pd.DataFrame):
        # ... (Omitted for brevity, unchanged)
        self.inventory_table.setRowCount(0)
        if df.empty:
            self._show_loading_state(self.inventory_table, "No inventory found.");
            self.total_balance_label.setText("Total Balance: 0.00 kg");
            return
        total_balance = df['current_balance'].sum();
        self.inventory_table.setRowCount(len(df))
        for i, row in df.iterrows():
            qty_item = QTableWidgetItem(f"{row.get('current_balance', 0.0):,.2f}");
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.inventory_table.setItem(i, 0, QTableWidgetItem(str(row.get('product_code', ''))));
            self.inventory_table.setItem(i, 1, QTableWidgetItem(str(row.get('lot_number', ''))))
            self.inventory_table.setItem(i, 2, qty_item);
            self.inventory_table.setItem(i, 3, QTableWidgetItem(str(row.get('bag_box_number', ''))));
            self.inventory_table.setItem(i, 4, QTableWidgetItem(str(row.get('location', 'N/A'))))
        is_filtered = any(
            [self.lot_number_input.text(), self.product_code_input.text(), self.fg_type_combo.currentText() != "All"]);
        prefix = f"Filtered Total ({len(df)} lots)" if is_filtered else f"Overall Total ({len(df)} lots)"
        self.total_balance_label.setText(f"{prefix}: {total_balance:,.2f} kg")

    def _format_excel_sheet(self, worksheet: Worksheet, df: pd.DataFrame):
        # ... (Omitted for brevity, unchanged)
        qty_col_name = next((col for col in df.columns if col.upper() == 'QTY'), None)
        if qty_col_name:
            qty_col_idx = df.columns.get_loc(qty_col_name) + 1;
            qty_col_letter = get_column_letter(qty_col_idx)
            for cell in worksheet[qty_col_letter][1:]: cell.number_format = '#,##0.00'
        for i, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(i);
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2;
            worksheet.column_dimensions[column_letter].width = max_length

    def _export_to_excel(self):
        # ... (Omitted for brevity, unchanged)
        if self.current_inventory_df.empty: QMessageBox.information(self, "Export Failed", "No data to export."); return
        default_filename = f"FG_Inventory_Report_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Report", default_filename, "Excel Files (*.xlsx)")
        if filepath:
            try:
                export_df = self.current_inventory_df.rename(
                    columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                             'location': 'LOCATION', 'bag_box_number': 'BAG/BOX_NUMBER'})
                final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'QTY', 'BAG/BOX_NUMBER', 'LOCATION'];
                df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
                df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy()
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    if not df_mb.empty: df_mb.to_excel(writer, index=False,
                                                       sheet_name='FG MB PASSED'); self._format_excel_sheet(
                        writer.sheets['FG MB PASSED'], df_mb)
                    if not df_dc.empty: df_dc.to_excel(writer, index=False,
                                                       sheet_name='FG DC PASSED'); self._format_excel_sheet(
                        writer.sheets['FG DC PASSED'], df_dc)
                QMessageBox.information(self, "Export Successful", f"Data exported to:\n{os.path.basename(filepath)}")
            except Exception:
                show_error_message(self, "Export Error", "Failed to save file.", traceback.format_exc())

    # <<< NEW: METHOD TO OPEN THE SETTINGS DIALOG >>>
    def _open_settings_dialog(self):
        """Creates and shows the settings dialog."""
        dialog = SettingsDialog(self)
        dialog.exec()

    # <<< MODIFIED: EMAIL HANDLER NOW USES QSETTINGS >>>
    def _export_and_email(self):
        if self.current_inventory_df.empty:
            QMessageBox.information(self, "Email Failed", "No data available to email.")
            return

        settings = QSettings("MyCompany", "FGInventoryApp")
        email_config = {
            "sender_email": settings.value("email/sender_email", ""),
            "sender_password": settings.value("email/sender_password", ""),
            "smtp_server": settings.value("email/smtp_server", ""),
            "smtp_port": settings.value("email/smtp_port", 0, type=int),
            "recipient_email": settings.value("email/recipient_email", "")
        }

        if not all([email_config["sender_email"], email_config["sender_password"], email_config["smtp_server"],
                    email_config["smtp_port"], email_config["recipient_email"]]):
            QMessageBox.warning(self, "Configuration Needed",
                                "Email settings are incomplete. Please use the 'Settings' button to configure them first.")
            self._open_settings_dialog()
            return

        try:
            export_df = self.current_inventory_df.rename(
                columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                         'location': 'LOCATION', 'bag_box_number': 'BAG/BOX_NUMBER'})
            final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'QTY', 'BAG/BOX_NUMBER', 'LOCATION']
            df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy()
            df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy()
            df_dict = {}
            if not df_mb.empty: df_dict['FG MB PASSED'] = df_mb
            if not df_dc.empty: df_dict['FG DC PASSED'] = df_dc

            if not df_dict:
                QMessageBox.warning(self, "Email Failed", "No data to include in the email after filtering.")
                return

            as_of_date_str = self.date_picker.date().toString("yyyy-MM-dd")
            email_config['subject'] = f"FG Inventory Report as of {as_of_date_str}"
            email_config['excel_fn'] = f"FG_Inventory_Report_{datetime.now():%Y-%m-%d}.xlsx"

            self.set_controls_enabled(False)
            QApplication.processEvents()

            send_email_with_excel(df_dict, email_config)

            QMessageBox.information(self, "Email Sent",
                                    f"The report has been successfully sent to {email_config['recipient_email']}.")

        except Exception as e:
            show_error_message(self, "Email Error", "Failed to send email. Please check your settings and connection.",
                               f"Error details:\n\n{traceback.format_exc()}")
        finally:
            self.set_controls_enabled(True)


# --- Main Application Window (unchanged) ---
class InventoryApp(QMainWindow):
    def __init__(self, engine: Engine):
        super().__init__()
        self.setWindowTitle("FG Inventory Computation and Export")
        self.setGeometry(100, 100, 1400, 900)
        self.setCentralWidget(GoodInventoryPage(engine))


# --- Database Connection Setup (unchanged) ---
def setup_postgres_engine() -> Engine | None:
    DB_CONFIG = {"host": "192.168.1.13", "port": 5432, "database": "dbfg", "username": "postgres", "password": "mbpi"}
    db_url = URL.create("postgresql+psycopg2", **DB_CONFIG)
    try:
        engine = create_engine(db_url, pool_recycle=3600, connect_args={'connect_timeout': 5})
        with engine.connect():
            pass
        return engine
    except Exception:
        error_msg = f"Failed to connect to PostgreSQL at {DB_CONFIG['host']}:{DB_CONFIG['port']}."
        detailed_text = traceback.format_exc()
        temp_app = QApplication.instance() or QApplication(sys.argv)
        show_error_message(None, "Database Connection Error", error_msg, detailed_text)
        return None


# --- Main Execution Block (unchanged) ---
if __name__ == "__main__":
    app = QApplication(sys.argv)
    db_engine = setup_postgres_engine()
    if db_engine:
        main_window = InventoryApp(db_engine)
        main_window.show()
        sys.exit(app.exec())
    else:
        sys.exit(1)