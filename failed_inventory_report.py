import pandas as pd
import os
import sys
import traceback
from datetime import datetime
import qtawesome as fa

# --- Email Imports ---
from email.message import EmailMessage
from io import BytesIO
import smtplib
import ssl

# --- PyQt6 Imports ---
from PyQt6.QtCore import Qt, QObject, pyqtSignal, QThread, QDate, QSize, QSettings, QTimer, QPoint
from PyQt6.QtGui import QColor, QAction
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
                             QHeaderView, QMessageBox, QHBoxLayout, QLabel,
                             QPushButton, QLineEdit, QAbstractItemView, QDateEdit,
                             QGroupBox, QFileDialog, QTabWidget, QGridLayout,
                             QProgressBar, QDialog, QFormLayout, QDialogButtonBox,
                             QSpinBox, QApplication, QMainWindow, QComboBox, QMenu)

# --- SQLAlchemy and OpenPyXL Imports ---
from sqlalchemy import text, create_engine
from sqlalchemy.engine import Engine
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- UI CONSTANTS ---
PRIMARY_ACCENT_COLOR = "#007bff"
DANGER_ACCENT_COLOR = "#dc3545"
NEUTRAL_COLOR = "#6c757d"
LIGHT_TEXT_COLOR = "#333333"
BACKGROUND_CONTENT_COLOR = "#f4f7fc"
INPUT_BACKGROUND_COLOR = "#ffffff"
GROUP_BOX_HEADER_COLOR = "#f4f7fc"
TABLE_SELECTION_COLOR = "#3a506b"
HEADER_AND_ICON_COLOR = "#3a506b"
TABLE_HEADER_TEXT_COLOR = "#4f4f4f"
ERROR_COLOR = "#dc3545"


# --- Helper Widgets and Functions ---
class UpperCaseLineEdit(QLineEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.textEdited.connect(self._to_upper)

    def _to_upper(self, text: str):
        if text and text != self.text().upper():
            self.blockSignals(True)
            self.setText(text.upper())
            self.blockSignals(False)


def show_error_message(parent, title, message, detailed_text=""):
    msg_box = QMessageBox(parent)
    msg_box.setIcon(QMessageBox.Icon.Critical)
    msg_box.setWindowTitle(title)
    msg_box.setText(f"<b>{message}</b>")
    if detailed_text:
        msg_box.setDetailedText(detailed_text)
    msg_box.exec()


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Email Settings")
        self.setMinimumWidth(450)
        self.settings = QSettings("MyCompany", "FGInventoryApp")
        main_layout = QVBoxLayout(self)
        sender_group = QGroupBox("Sender Details (SMTP)")
        sender_layout = QFormLayout(sender_group)
        self.sender_email_input = QLineEdit()
        self.sender_password_input = QLineEdit()
        self.sender_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.smtp_server_input = QLineEdit()
        self.smtp_port_input = QSpinBox()
        self.smtp_port_input.setRange(1, 65535)
        sender_layout.addRow("Sender Email:", self.sender_email_input)
        sender_layout.addRow("Password:", self.sender_password_input)
        sender_layout.addRow("SMTP Server:", self.smtp_server_input)
        sender_layout.addRow("SMTP Port:", self.smtp_port_input)
        main_layout.addWidget(sender_group)
        recipient_group = QGroupBox("Default Recipient")
        recipient_layout = QFormLayout(recipient_group)
        self.recipient_email_input = QLineEdit()
        self.recipient_email_input.setPlaceholderText("e.g., user1@example.com, user2@example.com")
        recipient_layout.addRow("Recipient Email(s):", self.recipient_email_input)
        main_layout.addWidget(recipient_group)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        main_layout.addWidget(button_box)
        self.load_settings()

    def load_settings(self):
        self.sender_email_input.setText(self.settings.value("email/sender_email", ""))
        if self.settings.value("email/sender_password"):
            self.sender_password_input.setPlaceholderText("Password is saved. Enter new to change.")
        self.smtp_server_input.setText(self.settings.value("email/smtp_server", ""))
        self.smtp_port_input.setValue(int(self.settings.value("email/smtp_port", 587)))
        self.recipient_email_input.setText(self.settings.value("email/recipient_email", ""))

    def accept(self):
        self.settings.setValue("email/sender_email", self.sender_email_input.text().strip())
        if self.sender_password_input.text():
            self.settings.setValue("email/sender_password", self.sender_password_input.text())
        self.settings.setValue("email/smtp_server", self.smtp_server_input.text().strip())
        self.settings.setValue("email/smtp_port", self.smtp_port_input.value())
        self.settings.setValue("email/recipient_email", self.recipient_email_input.text().strip())
        QMessageBox.information(self, "Success", "Settings have been saved.")
        super().accept()


class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Report Date Range")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        self.start_date_edit = QDateEdit(calendarPopup=True, displayFormat="yyyy-MM-dd")
        self.end_date_edit = QDateEdit(calendarPopup=True, displayFormat="yyyy-MM-dd")
        today = QDate.currentDate()
        first_day_of_month = QDate(today.year(), today.month(), 1)
        self.start_date_edit.setDate(first_day_of_month)
        self.end_date_edit.setDate(today)
        layout.addRow("Start Date:", self.start_date_edit)
        layout.addRow("End Date:", self.end_date_edit)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_dates(self):
        return self.start_date_edit.date(), self.end_date_edit.date()


def send_email_with_excel(attachments, email_config):
    sender_email = email_config.get('sender_email')
    sender_password = email_config.get('sender_password')
    smtp_server = email_config.get('smtp_server')
    smtp_port = email_config.get('smtp_port')
    recipient_email = email_config.get('recipient_email')
    subject = email_config.get('subject')
    if not all([sender_email, sender_password, smtp_server, smtp_port, recipient_email]):
        raise ValueError("Incomplete email settings. Please configure them in the Settings menu.")
    recipients = [email.strip() for email in recipient_email.split(',')]
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipients)
    body = f'Please find the attached Excel report(s): {", ".join(attachments.keys())}.\n\nThis is an automated message.'
    msg.set_content(body)
    for filename, df_dict in attachments.items():
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for sheet_name, df in df_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        excel_buffer.seek(0)
        msg.add_attachment(excel_buffer.read(), maintype='application',
                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)
    context = ssl.create_default_context()
    smtp_port_int = int(smtp_port)
    if smtp_port_int == 465:
        with smtplib.SMTP_SSL(smtp_server, smtp_port_int, context=context) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg, from_addr=sender_email, to_addrs=recipients)
    else:
        with smtplib.SMTP(smtp_server, smtp_port_int) as server:
            server.starttls(context=context)
            server.login(sender_email, sender_password)
            server.send_message(msg, from_addr=sender_email, to_addrs=recipients)


class EmailWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str, str)
    success = pyqtSignal(str)

    def __init__(self, attachments, email_config, log_func, log_action, recipient):
        super().__init__()
        self.attachments = attachments
        self.email_config = email_config
        self.log_audit_trail = log_func
        self.log_action = log_action
        self.recipient = recipient

    def run(self):
        try:
            send_email_with_excel(self.attachments, self.email_config)
            self.log_audit_trail(self.log_action, f"User emailed report to '{self.recipient}'.")
            self.success.emit(f"The report has been successfully sent to {self.recipient}.")
        except ValueError as e:
            self.error.emit("Email Configuration Error", str(e))
        except Exception:
            self.error.emit("Network/SMTP Error", traceback.format_exc())
        finally:
            self.finished.emit()


class FailedInventoryWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, product_filter: str, lot_filter: str, as_of_date: str):
        super().__init__()
        self.engine = engine
        self.product_filter = product_filter
        self.lot_filter = lot_filter
        self.as_of_date = as_of_date

    def run(self):
        try:
            params = {'as_of_date': self.as_of_date}
            product_filter_raw_clause = ""
            if self.product_filter:
                product_filter_raw_clause = "AND UPPER(TRIM(product_code)) LIKE :product_search"
                params['product_search'] = f"%{self.product_filter}%"
            lot_filter_raw_clause = ""
            if self.lot_filter:
                lot_filter_raw_clause = "AND UPPER(TRIM(lot_number)) LIKE :lot_search"
                params['lot_search'] = f"%{self.lot_filter}%"
            query_str = f"""
                WITH all_failed_tx AS (
                    SELECT
                        UPPER(TRIM(product_code)) AS product_code, UPPER(TRIM(lot_number)) AS lot_number,
                        COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in, 0.0 AS quantity_out,
                        production_date AS transaction_date, UPPER(TRIM(location)) as location,
                        COALESCE(UPPER(TRIM(bag_number)), UPPER(TRIM(box_number)), '') as bag_box_number,
                        'BEGINV' as transaction_type
                    FROM beg_invfailed1
                    WHERE product_code IS NOT NULL AND TRIM(product_code) <> '' AND lot_number IS NOT NULL AND TRIM(lot_number) <> ''
                      {product_filter_raw_clause} {lot_filter_raw_clause}
                    UNION ALL
                    SELECT
                        UPPER(TRIM(ft.product_code)) AS product_code, UPPER(TRIM(ft.lot_number)) AS lot_number,
                        COALESCE(CAST(ft.quantity_in AS NUMERIC), 0) AS quantity_in,
                        COALESCE(CAST(ft.quantity_out AS NUMERIC), 0) AS quantity_out,
                        ft.transaction_date, UPPER(TRIM(ft.warehouse)) as location,
                        COALESCE(qcf.bag_no, '') as bag_box_number, ft.transaction_type
                    FROM failed_transactions ft
                    LEFT JOIN qcf_endorsements_primary qcf ON ft.source_ref_no = qcf.system_ref_no
                    WHERE ft.product_code IS NOT NULL AND TRIM(ft.product_code) <> '' AND ft.lot_number IS NOT NULL AND TRIM(ft.lot_number) <> ''
                      AND ft.transaction_date <= :as_of_date
                      {product_filter_raw_clause.replace('product_code', 'ft.product_code')}
                      {lot_filter_raw_clause.replace('lot_number', 'ft.lot_number')}
                ),
                ranked_lot_details AS (
                    SELECT lot_number, location, bag_box_number,
                        ROW_NUMBER() OVER (
                            PARTITION BY lot_number ORDER BY
                                CASE WHEN bag_box_number IS NOT NULL AND bag_box_number <> '' THEN 0 ELSE 1 END,
                                transaction_date DESC NULLS LAST,
                                CASE WHEN transaction_type = 'BEGINV' THEN 1 ELSE 0 END
                        ) as rn
                    FROM all_failed_tx
                ),
                lot_summary AS (
                    SELECT lot_number, MAX(product_code) AS product_code,
                        CASE WHEN MAX(product_code) LIKE '%-%' THEN 'DC' ELSE 'MB' END AS fg_type,
                        COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0) AS current_balance
                    FROM all_failed_tx
                    GROUP BY lot_number
                    HAVING (COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0)) != 0
                )
                SELECT s.product_code, s.lot_number, s.current_balance, d.location, d.bag_box_number, s.fg_type
                FROM lot_summary s
                JOIN ranked_lot_details d ON s.lot_number = d.lot_number AND d.rn = 1
                WHERE d.location IS NOT NULL
                ORDER BY s.product_code, s.lot_number;
            """
            with self.engine.connect() as conn:
                results = conn.execute(text(query_str), params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame(
                columns=['product_code', 'lot_number', 'current_balance', 'location', 'bag_box_number', 'fg_type'])
            if not df.empty:
                df['current_balance'] = pd.to_numeric(df['current_balance'])
            self.finished.emit(df)
        except Exception:
            self.error.emit("Database query failed.", traceback.format_exc())


class FailedTransactionHistoryWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, lot_number: str, as_of_date: str):
        super().__init__()
        self.engine = engine
        self.lot_number = lot_number
        self.as_of_date = as_of_date

    def run(self):
        try:
            query = text("""
                SELECT
                    production_date AS transaction_date,
                    'BEGINV' AS transaction_type,
                    COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in,
                    0.0 AS quantity_out,
                    location,
                    'N/A' as source_ref_no
                FROM beg_invfailed1
                WHERE UPPER(TRIM(lot_number)) = :lot_number

                UNION ALL

                SELECT
                    ft.transaction_date,
                    ft.transaction_type,
                    COALESCE(CAST(ft.quantity_in AS NUMERIC), 0) AS quantity_in,
                    COALESCE(CAST(ft.quantity_out AS NUMERIC), 0) AS quantity_out,
                    ft.warehouse AS location,
                    ft.source_ref_no
                FROM failed_transactions ft
                WHERE UPPER(TRIM(ft.lot_number)) = :lot_number
                  AND ft.transaction_date <= :as_of_date

                ORDER BY transaction_date;
            """)
            params = {'lot_number': self.lot_number, 'as_of_date': self.as_of_date}
            with self.engine.connect() as conn:
                results = conn.execute(query, params).mappings().all()

            df = pd.DataFrame(results) if results else pd.DataFrame()
            self.finished.emit(df)
        except Exception:
            self.error.emit("Failed Lot History Query Failed", traceback.format_exc())


class TransactionHistoryDialog(QDialog):
    def __init__(self, worker_class, engine: Engine, lot_number: str, as_of_date: str, parent=None):
        super().__init__(parent)
        self.worker_class = worker_class
        self.engine = engine
        self.lot_number = lot_number
        self.as_of_date = as_of_date
        self.worker_thread = None
        self.worker = None

        self.setWindowTitle(f"Transaction History: {self.lot_number}")
        self.setMinimumSize(800, 500)
        self.setStyleSheet(parent.styleSheet() if parent else "")

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"<b>History for Lot:</b> {self.lot_number}", objectName="PageHeader"))

        self.history_table = QTableWidget()
        self.setup_table()
        layout.addWidget(self.history_table, 1)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.fetch_transactions()

    def setup_table(self):
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels([
            "Date", "Type", "Ref No.", "Qty In", "Qty Out", "Running Balance"
        ])
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

    def fetch_transactions(self):
        self._show_loading_state("Fetching transaction history...")
        self.worker_thread = QThread()
        self.worker = self.worker_class(self.engine, self.lot_number, self.as_of_date)
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(self._display_history)
        self.worker.error.connect(self._on_fetch_error)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker_thread.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker_thread.start()

    def _display_history(self, df: pd.DataFrame):
        self.history_table.setRowCount(0)
        if df.empty:
            self._show_loading_state("No transactions found for this lot.")
            return

        df['running_balance'] = (df['quantity_in'].astype(float) - df['quantity_out'].astype(float)).cumsum()
        self.history_table.setRowCount(len(df))
        for i, row in df.iterrows():
            date_str = row['transaction_date'].strftime('%Y-%m-%d') if pd.notna(row['transaction_date']) and row[
                'transaction_type'] != 'BEGINV' else 'Beginning'
            self.history_table.setItem(i, 0, QTableWidgetItem(date_str))
            self.history_table.setItem(i, 1, QTableWidgetItem(str(row.get('transaction_type', ''))))
            self.history_table.setItem(i, 2, QTableWidgetItem(str(row.get('source_ref_no', ''))))
            for col_idx, col_name in enumerate(['quantity_in', 'quantity_out', 'running_balance'], 3):
                item = QTableWidgetItem(f"{row.get(col_name, 0.0):,.2f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.history_table.setItem(i, col_idx, item)

    def _on_fetch_error(self, title, message):
        show_error_message(self, title, "Could not fetch lot history.", message)
        self._show_loading_state("Error fetching data.")

    def _show_loading_state(self, message: str):
        self.history_table.setRowCount(1)
        self.history_table.setSpan(0, 0, 1, self.history_table.columnCount())
        loading_item = QTableWidgetItem(message)
        loading_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.history_table.setItem(0, 0, loading_item)


class FailedMonthlyReportWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, start_date: str, end_date: str):
        super().__init__()
        self.engine = engine
        self.start_date = start_date
        self.end_date = end_date

    def run(self):
        try:
            params = {'start_date': self.start_date, 'end_date': self.end_date}
            query_str = f"""
                WITH period_failed_tx AS (
                    SELECT
                        UPPER(TRIM(product_code)) AS product_code, UPPER(TRIM(lot_number)) AS lot_number,
                        COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in, 0.0 AS quantity_out,
                        production_date AS transaction_date, UPPER(TRIM(location)) as location,
                        COALESCE(UPPER(TRIM(bag_number)), UPPER(TRIM(box_number)), '') as bag_box_number
                    FROM beg_invfailed1
                    WHERE product_code IS NOT NULL AND TRIM(product_code) <> '' AND lot_number IS NOT NULL AND TRIM(lot_number) <> ''
                      AND production_date BETWEEN :start_date AND :end_date
                    UNION ALL
                    SELECT
                        UPPER(TRIM(ft.product_code)) AS product_code, UPPER(TRIM(ft.lot_number)) AS lot_number,
                        COALESCE(CAST(ft.quantity_in AS NUMERIC), 0) AS quantity_in,
                        COALESCE(CAST(ft.quantity_out AS NUMERIC), 0) AS quantity_out,
                        ft.transaction_date, UPPER(TRIM(ft.warehouse)) as location,
                        COALESCE(qcf.bag_no, '') as bag_box_number
                    FROM failed_transactions ft
                    LEFT JOIN qcf_endorsements_primary qcf ON ft.source_ref_no = qcf.system_ref_no
                    WHERE ft.product_code IS NOT NULL AND TRIM(ft.product_code) <> '' AND ft.lot_number IS NOT NULL AND TRIM(ft.lot_number) <> ''
                      AND ft.transaction_date BETWEEN :start_date AND :end_date
                ),
                lot_details AS (
                    SELECT DISTINCT ON (lot_number) lot_number, location, bag_box_number
                    FROM period_failed_tx
                    ORDER BY lot_number, transaction_date DESC
                ),
                lot_summary AS (
                    SELECT
                        lot_number, MAX(product_code) AS product_code,
                        COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0) AS net_quantity
                    FROM period_failed_tx
                    GROUP BY lot_number
                    HAVING (COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0)) != 0
                )
                SELECT s.product_code, s.lot_number, s.net_quantity, d.location, d.bag_box_number
                FROM lot_summary s
                LEFT JOIN lot_details d ON s.lot_number = d.lot_number
                ORDER BY s.product_code, s.lot_number;
            """
            with self.engine.connect() as conn:
                results = conn.execute(text(query_str), params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame(
                columns=['product_code', 'lot_number', 'net_quantity', 'location', 'bag_box_number'])
            if not df.empty:
                df['net_quantity'] = pd.to_numeric(df['net_quantity'])
            self.finished.emit(df)
        except Exception:
            self.error.emit("Failed Monthly Report query failed.", traceback.format_exc())


class FailedDashboardWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data_df = pd.DataFrame()
        self.init_ui()
        self.setStyleSheet(self._get_dashboard_styles())

    def _get_dashboard_styles(self) -> str:
        return f"QGroupBox {{ border: 1px solid #e0e5eb; border-radius: 8px; margin-top: 12px; background-color: {INPUT_BACKGROUND_COLOR}; }} QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top left; padding: 2px 10px; background-color: {GROUP_BOX_HEADER_COLOR}; border: 1px solid #e0e5eb; border-bottom: none; border-top-left-radius: 8px; border-top-right-radius: 8px; font-weight: bold; color: {HEADER_AND_ICON_COLOR}; }} QLabel#TitleLabel {{ font-size: 10pt; color: {HEADER_AND_ICON_COLOR}; background-color: transparent; }} QLabel#ValueLabel {{ font-size: 16pt; font-weight: bold; color: {DANGER_ACCENT_COLOR}; background-color: transparent; }} QTableWidget {{ border: 1px solid #e0e5eb; background-color: {INPUT_BACKGROUND_COLOR}; gridline-color: #f0f3f8; }} QTableWidget::item:selected {{ background-color: {TABLE_SELECTION_COLOR}; color: white; }} QHeaderView::section {{ background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR}; }} QProgressBar::chunk {{ background-color: {DANGER_ACCENT_COLOR}; border-radius: 4px; }}"

    def _create_summary_box(self, title: str, initial_value: str) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(15, 10, 15, 10)
        title_label = QLabel(title, objectName="TitleLabel")
        value_label = QLabel(initial_value, objectName="ValueLabel", alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label, alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_label, alignment=Qt.AlignmentFlag.AlignCenter)
        widget.setStyleSheet(
            f"background-color: {INPUT_BACKGROUND_COLOR}; border: 1px solid #d1d9e6; border-radius: 8px;")
        return widget

    def _create_contribution_table(self) -> QTableWidget:
        table = QTableWidget(columnCount=4)
        table.setHorizontalHeaderLabels(["Product Code", "Balance (kg)", "% Share", ""])
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.verticalHeader().setVisible(False)
        table.setShowGrid(True)
        table.setAlternatingRowColors(True)
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        return table

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(15)
        grid_layout = QGridLayout()
        grid_layout.setRowStretch(1, 1)
        grid_layout.setColumnStretch(0, 1)
        grid_layout.setColumnStretch(1, 1)
        summary_group = QGroupBox("Overall Failed Metrics")
        summary_layout = QHBoxLayout(summary_group)
        self.total_lots_label = self._create_summary_box("Total Failed Lots:", "0")
        self.total_products_label = self._create_summary_box("Unique Failed Products:", "0")
        self.overall_balance_label = self._create_summary_box("Total Failed Balance (kg):", "0.00")
        summary_layout.addWidget(self.total_lots_label)
        summary_layout.addWidget(self.total_products_label)
        summary_layout.addWidget(self.overall_balance_label)
        grid_layout.addWidget(summary_group, 0, 0, 1, 2)
        self.contribution_table = self._create_contribution_table()
        contribution_group = QGroupBox("Top 10 Failed Products by Mass")
        vbox_contribution = QVBoxLayout(contribution_group)
        vbox_contribution.addWidget(self.contribution_table)
        grid_layout.addWidget(contribution_group, 1, 0)
        main_layout.addLayout(grid_layout)

    def update_dashboard(self, df: pd.DataFrame):
        is_empty = df.empty
        positive_balance_df = df[df['current_balance'] > 0] if not is_empty else pd.DataFrame()

        overall_balance = df['current_balance'].sum() if not is_empty else 0.0

        self.total_lots_label.findChild(QLabel, "ValueLabel").setText("0" if is_empty else f"{len(df):,}")
        self.total_products_label.findChild(QLabel, "ValueLabel").setText(
            "0" if is_empty else f"{df['product_code'].nunique():,}")
        self.overall_balance_label.findChild(QLabel, "ValueLabel").setText(f"{overall_balance:,.2f}")

        self.contribution_table.setRowCount(0)
        if positive_balance_df.empty:
            return

        positive_total = positive_balance_df['current_balance'].sum()
        if positive_total <= 0:
            return

        summary = positive_balance_df.groupby('product_code')['current_balance'].sum().nlargest(10).reset_index()
        summary['percentage'] = (summary['current_balance'] / positive_total) * 100
        self.contribution_table.setRowCount(len(summary))
        for i, row in summary.iterrows():
            self.contribution_table.setItem(i, 0, QTableWidgetItem(str(row['product_code'])))
            balance_item = QTableWidgetItem(f"{row['current_balance']:,.2f}")
            balance_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.contribution_table.setItem(i, 1, balance_item)
            percent_item = QTableWidgetItem(f"{row['percentage']:.1f}%")
            percent_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.contribution_table.setItem(i, 2, percent_item)
            progress_bar = QProgressBar(maximum=100, value=int(row['percentage']), textVisible=False)
            self.contribution_table.setCellWidget(i, 3, progress_bar)


class FailedInventoryReportPage(QWidget):
    def __init__(self, engine: Engine, username: str, log_audit_trail_func, good_inventory_page=None):
        super().__init__()
        self.engine = engine;
        self.username = username;
        self.log_audit_trail = log_audit_trail_func;
        self.good_inventory_page = good_inventory_page
        self.inventory_thread: QThread | None = None;
        self.inventory_worker: FailedInventoryWorker | None = None
        self.email_thread: QThread | None = None;
        self.email_worker: EmailWorker | None = None
        self.current_inventory_df = pd.DataFrame();
        self.is_calculating = False
        self.failed_monthly_report_thread: QThread | None = None;
        self.failed_monthly_report_worker: FailedMonthlyReportWorker | None = None
        self.init_ui()
        self.setStyleSheet(self._get_styles())
        QTimer.singleShot(100, self._start_inventory_calculation)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        header_layout = QHBoxLayout();
        header_layout.setContentsMargins(0, 0, 0, 0)
        icon_label = QLabel();
        icon_label.setPixmap(fa.icon('fa5s.times-circle', color=DANGER_ACCENT_COLOR).pixmap(QSize(28, 28)))
        header_layout.addWidget(icon_label);
        header_layout.addWidget(QLabel("FG Inventory Report (Failed)", objectName="PageHeader"));
        header_layout.addStretch()
        main_layout.addLayout(header_layout)
        instruction_group = QGroupBox("Instructions");
        instruction_layout = QVBoxLayout(instruction_group)
        instruction_label = QLabel(
            "This page calculates the inventory balance for FAILED items up to the selected date. Use the filters below to refine the results.")
        instruction_label.setStyleSheet("font-style: italic; color: #555;");
        instruction_label.setWordWrap(True)
        instruction_layout.addWidget(instruction_label);
        main_layout.addWidget(instruction_group)
        main_layout.addWidget(self._create_filter_controls())
        self.tab_widget = QTabWidget();
        self.tab_widget.setIconSize(QSize(16, 16));
        self.inventory_tab = QWidget();
        inventory_layout = QVBoxLayout(self.inventory_tab)

        self.inventory_table = self._create_inventory_table()
        self.inventory_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.inventory_table.customContextMenuRequested.connect(self.show_table_context_menu)

        self.total_balance_label = QLabel("Total Balance: 0.00 kg", objectName="TotalBalanceLabel",
                                          alignment=Qt.AlignmentFlag.AlignRight)
        inventory_layout.addWidget(self.inventory_table);
        inventory_layout.addWidget(self.total_balance_label)
        self.tab_widget.addTab(self.inventory_tab, fa.icon('fa5.list-alt', color=DANGER_ACCENT_COLOR),
                               "Failed Inventory Details")
        self.dashboard_widget = FailedDashboardWidget()
        self.tab_widget.addTab(self.dashboard_widget, fa.icon('fa5s.chart-area', color=DANGER_ACCENT_COLOR),
                               "Dashboard")
        main_layout.addWidget(self.tab_widget, 1)

    def show_table_context_menu(self, pos: QPoint):
        item = self.inventory_table.itemAt(pos)
        if not item: return

        row_index = item.row()
        lot_item = self.inventory_table.item(row_index, 1)
        if not lot_item: return

        lot_number = lot_item.text()
        context_menu = QMenu(self)
        track_action = QAction(fa.icon('fa5s.history', color=HEADER_AND_ICON_COLOR), f"Track Lot: {lot_number}", self)
        track_action.triggered.connect(lambda: self.track_lot_transactions(lot_number))
        context_menu.addAction(track_action)
        context_menu.exec(self.inventory_table.mapToGlobal(pos))

    def track_lot_transactions(self, lot_number: str):
        as_of_date_str = self.date_picker.date().toString(Qt.DateFormat.ISODate)
        dialog = TransactionHistoryDialog(FailedTransactionHistoryWorker, self.engine, lot_number, as_of_date_str, self)
        dialog.exec()

    def _display_inventory(self, df: pd.DataFrame):
        self.inventory_table.setRowCount(0)
        if df.empty:
            is_filtered = bool(self.product_code_input.text() or self.lot_number_input.text())
            message = "No matching inventory found based on filters." if is_filtered else "No failed inventory found."
            self._show_loading_state(self.inventory_table, message)
            self.total_balance_label.setText("Total Balance: 0.00 kg")
            return
        total_balance = df['current_balance'].sum()
        self.inventory_table.setRowCount(len(df))
        for i, row in df.iterrows():
            current_balance = row.get('current_balance', 0.0)
            qty_item = QTableWidgetItem(f"{current_balance:,.2f}")
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

            if current_balance < 0:
                qty_item.setForeground(QColor(ERROR_COLOR))

            self.inventory_table.setItem(i, 0, QTableWidgetItem(str(row.get('product_code', ''))))
            self.inventory_table.setItem(i, 1, QTableWidgetItem(str(row.get('lot_number', ''))))
            self.inventory_table.setItem(i, 2, qty_item)
            self.inventory_table.setItem(i, 3, QTableWidgetItem(str(row.get('bag_box_number', ''))))
            self.inventory_table.setItem(i, 4, QTableWidgetItem(str(row.get('location', 'N/A'))))
        is_filtered = bool(self.product_code_input.text() or self.lot_number_input.text())
        prefix = f"Filtered Total ({len(df)} lots)" if is_filtered else f"Overall Total ({len(df)} lots)"
        self.total_balance_label.setText(f"{prefix}: {total_balance:,.2f} kg")

    def refresh_page(self):
        self._start_inventory_calculation()

    def _get_styles(self) -> str:
        return f"QWidget{{background-color:{BACKGROUND_CONTENT_COLOR}; color:{LIGHT_TEXT_COLOR}; font-family:'Segoe UI',Arial,sans-serif;}} QGroupBox{{border:1px solid #e0e5eb; border-radius:8px; margin-top:12px; background-color:{INPUT_BACKGROUND_COLOR};}} QGroupBox::title{{subcontrol-origin:margin; subcontrol-position:top left; padding:2px 10px; background-color:{GROUP_BOX_HEADER_COLOR}; border:1px solid #e0e5eb; border-bottom:1px solid {INPUT_BACKGROUND_COLOR}; border-top-left-radius:8px; border-top-right-radius:8px; font-weight:bold; color:#4f4f4f;}} QGroupBox QLabel{{background-color: transparent;}} QLabel#PageHeader{{font-size:15pt; font-weight:bold; color:{DANGER_ACCENT_COLOR}; background-color:transparent;}} QLineEdit, QDateEdit, QComboBox{{border:1px solid #d1d9e6; padding:8px; border-radius:5px; background-color:{INPUT_BACKGROUND_COLOR};}} QLineEdit:focus, QDateEdit:focus, QComboBox:focus{{border:1px solid {PRIMARY_ACCENT_COLOR};}} QPushButton{{border:1px solid #d1d9e6; padding:8px 15px; border-radius:6px; font-weight:bold; background-color:{INPUT_BACKGROUND_COLOR};}} QPushButton:hover{{background-color:#f0f3f8;}} QPushButton#PrimaryButton{{color:{DANGER_ACCENT_COLOR}; border:1px solid {DANGER_ACCENT_COLOR};}} QPushButton#PrimaryButton:hover{{background-color:#fef0f0;}} QTabWidget::pane{{border:1px solid #e0e5eb; border-radius:8px; background-color:{INPUT_BACKGROUND_COLOR}; padding:10px; margin-top:-1px;}} QTabBar::tab{{background:#e9eff7; color:{NEUTRAL_COLOR}; padding:8px 15px; border:1px solid #e0e5eb; border-bottom:none; border-top-left-radius:6px; border-top-right-radius:6px;}} QTabBar::tab:selected{{color:{DANGER_ACCENT_COLOR}; background:{INPUT_BACKGROUND_COLOR}; border-bottom-color:{INPUT_BACKGROUND_COLOR}; font-weight:bold;}} QTableWidget{{border:1px solid #e0e5eb; background-color:{INPUT_BACKGROUND_COLOR}; selection-behavior:SelectRows; gridline-color: #f0f3f8;}} QTableWidget::item:hover{{background-color: transparent;}} QTableWidget::item:selected{{background-color:{TABLE_SELECTION_COLOR}; color:white;}} QHeaderView::section{{background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR};}} QLabel#TotalBalanceLabel{{background-color:{INPUT_BACKGROUND_COLOR}; color:{DANGER_ACCENT_COLOR}; padding:10px; border-radius:6px; border:1px solid #e0e5eb; font-weight:bold;}}"

    def _create_filter_controls(self) -> QGroupBox:
        group = QGroupBox("Filters & Actions");
        layout = QHBoxLayout(group);
        layout.addWidget(QLabel("As Of Date:"));
        self.date_picker = QDateEdit(calendarPopup=True, date=QDate.currentDate(), displayFormat="yyyy-MM-dd");
        layout.addWidget(self.date_picker);
        layout.addSpacing(10);
        layout.addWidget(QLabel("Product:"));
        self.product_code_input = UpperCaseLineEdit(placeholderText="Filter by Product Code");
        layout.addWidget(self.product_code_input, 1);
        layout.addSpacing(10);
        layout.addWidget(QLabel("Lot:"));
        self.lot_number_input = UpperCaseLineEdit(placeholderText="Filter by Lot Number");
        layout.addWidget(self.lot_number_input, 1);
        layout.addStretch();
        self.refresh_button = QPushButton("Refresh", objectName="PrimaryButton",
                                          icon=fa.icon('fa5s.sync-alt', color=DANGER_ACCENT_COLOR));
        self.export_button = QPushButton("Export to Excel", objectName="PrimaryButton",
                                         icon=fa.icon('fa5s.file-excel', color=DANGER_ACCENT_COLOR));
        self.email_button = QPushButton("Email (Failed Only)", objectName="PrimaryButton",
                                        icon=fa.icon('fa5s.paper-plane', color=DANGER_ACCENT_COLOR));
        self.email_combined_button = QPushButton("Email Combined Report", objectName="PrimaryButton",
                                                 icon=fa.icon('fa5s.envelope-open-text', color=DANGER_ACCENT_COLOR));
        self.monthly_report_button = QPushButton("Monthly Report", objectName="PrimaryButton",
                                                 icon=fa.icon('fa5s.calendar-alt', color=DANGER_ACCENT_COLOR));
        self.settings_button = QPushButton("Settings", objectName="PrimaryButton",
                                           icon=fa.icon('fa5s.cog', color=DANGER_ACCENT_COLOR));
        layout.addWidget(self.refresh_button);
        layout.addWidget(self.export_button);
        layout.addWidget(self.email_button);
        layout.addWidget(self.email_combined_button);
        layout.addWidget(self.monthly_report_button);
        layout.addWidget(self.settings_button);
        self.refresh_button.clicked.connect(self._start_inventory_calculation);
        self.product_code_input.returnPressed.connect(self._start_inventory_calculation);
        self.lot_number_input.returnPressed.connect(self._start_inventory_calculation);
        self.date_picker.dateChanged.connect(self._start_inventory_calculation);
        self.export_button.clicked.connect(self._export_to_excel);
        self.email_button.clicked.connect(self._export_and_email);
        self.email_combined_button.clicked.connect(self._export_and_email_combined_report);
        self.monthly_report_button.clicked.connect(self._handle_failed_monthly_report_request);
        self.settings_button.clicked.connect(self._open_settings_dialog);
        return group

    def _create_inventory_table(self) -> QTableWidget:
        table = QTableWidget(columnCount=5);
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers);
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows);
        table.verticalHeader().setVisible(False);
        table.setAlternatingRowColors(True);
        table.setHorizontalHeaderLabels(["PRODUCT", "LOT NUMBER", "QTY (kg)", "BAG/BOX NO.", "LOCATION"]);
        header = table.horizontalHeader();
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents);
        return table

    def _cleanup_thread_state(self):
        self.is_calculating = False; self.inventory_thread = None; self.inventory_worker = None; self.set_controls_enabled(
            True)

    def _start_inventory_calculation(self):
        if self.is_calculating: return
        self.is_calculating = True;
        self.set_controls_enabled(False);
        self._show_loading_state(self.inventory_table, "Calculating failed inventory balance...");
        self.dashboard_widget.update_dashboard(pd.DataFrame());
        QApplication.processEvents()
        date_str = self.date_picker.date().toString(Qt.DateFormat.ISODate);
        product_filter_clean = self.product_code_input.text().strip().upper();
        lot_filter_clean = self.lot_number_input.text().strip().upper();
        filters = f"Date: {date_str}, Prod: '{product_filter_clean}', Lot: '{lot_filter_clean}'"
        self.log_audit_trail("CALCULATE_FAILED_INVENTORY", f"User calculated failed inventory with filters: {filters}");
        self.inventory_thread = QThread();
        self.inventory_worker = FailedInventoryWorker(engine=self.engine, product_filter=product_filter_clean,
                                                      lot_filter=lot_filter_clean, as_of_date=date_str)
        self.inventory_worker.moveToThread(self.inventory_thread);
        self.inventory_thread.started.connect(self.inventory_worker.run);
        self.inventory_worker.finished.connect(self._on_inventory_finished);
        self.inventory_worker.error.connect(self._on_calculation_error);
        self.inventory_worker.finished.connect(self.inventory_thread.quit);
        self.inventory_worker.error.connect(self.inventory_thread.quit);
        self.inventory_thread.finished.connect(lambda: QTimer.singleShot(100, self._cleanup_thread_state));
        self.inventory_thread.finished.connect(self.inventory_worker.deleteLater);
        self.inventory_thread.finished.connect(self.inventory_thread.deleteLater);
        self.inventory_thread.start()

    def set_controls_enabled(self, enabled: bool):
        for widget in [self.refresh_button, self.date_picker, self.settings_button, self.product_code_input,
                       self.lot_number_input, self.export_button, self.email_button, self.email_combined_button,
                       self.monthly_report_button]: widget.setEnabled(enabled)
        if enabled and self.current_inventory_df.empty: self.export_button.setEnabled(
            False); self.email_button.setEnabled(False)
        if self.good_inventory_page and hasattr(self.good_inventory_page,
                                                'set_controls_enabled'): self.good_inventory_page.set_controls_enabled(
            enabled)

    def _on_inventory_finished(self, df: pd.DataFrame):
        try:
            self.current_inventory_df = df.copy(); self._display_inventory(df); self.dashboard_widget.update_dashboard(
                df)
        finally:
            self.set_controls_enabled(True)

    def _on_calculation_error(self, error_message: str, detailed_traceback: str):
        try:
            show_error_message(self, "Calculation Error", error_message, detailed_traceback); self._show_loading_state(
                self.inventory_table, "Error during calculation."); self.dashboard_widget.update_dashboard(
                pd.DataFrame())
        finally:
            self.set_controls_enabled(True)

    def _on_email_error(self, title: str, details: str):
        self.set_controls_enabled(True); show_error_message(self, title,
                                                            "Failed to send email. Please check your settings and connection.",
                                                            details)

    def _on_email_success(self, message: str):
        self.set_controls_enabled(True); QMessageBox.information(self, "Email Sent", message)

    def _show_loading_state(self, table: QTableWidget, message: str):
        table.setRowCount(1); table.setSpan(0, 0, 1, table.columnCount()); loading_item = QTableWidgetItem(
            message); loading_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter); table.setItem(0, 0, loading_item)

    def _format_excel_sheet(self, worksheet: Worksheet, df: pd.DataFrame):
        qty_col_name = next((col for col in df.columns if 'QTY' in col.upper() or 'BALANCE' in col.upper()), None)
        if qty_col_name and not df.empty: qty_col_idx = df.columns.get_loc(
            qty_col_name) + 1; qty_col_letter = get_column_letter(qty_col_idx); [cell.number_format for cell in
                                                                                 worksheet[qty_col_letter][1:] if
                                                                                 setattr(cell, 'number_format',
                                                                                         '#,##0.00')]
        for i, col in enumerate(df.columns, 1):
            if not df.empty: column_letter = get_column_letter(i); max_length = max(df[col].astype(str).map(len).max(),
                                                                                    len(col)) + 2;
            worksheet.column_dimensions[column_letter].width = max_length

    def _export_to_excel(self):
        if self.current_inventory_df.empty: QMessageBox.information(self, "Export Failed", "No data to export."); return
        default_filename = f"FG_Inventory_Report_FAILED_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Report", default_filename, "Excel Files (*.xlsx)")
        if filepath:
            try:
                export_df = self.current_inventory_df.rename(
                    columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                             'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
                final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
                df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
                df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy()
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    if not df_mb.empty: df_mb.to_excel(writer, index=False,
                                                       sheet_name='FG FAILED MB'); self._format_excel_sheet(
                        writer.sheets['FG FAILED MB'], df_mb)
                    if not df_dc.empty: df_dc.to_excel(writer, index=False,
                                                       sheet_name='FG FAILED DC'); self._format_excel_sheet(
                        writer.sheets['FG FAILED DC'], df_dc)
                filename = os.path.basename(filepath);
                self.log_audit_trail("EXPORT_FAILED_INVENTORY",
                                     f"User exported failed inventory report to '{filename}'.");
                QMessageBox.information(self, "Export Successful", f"Data exported to:\n{filename}")
            except Exception:
                show_error_message(self, "Export Error", "Failed to save file.", traceback.format_exc())

    def _open_settings_dialog(self):
        dialog = SettingsDialog(self); dialog.exec()

    def _handle_failed_monthly_report_request(self):
        if self.is_calculating: return
        dialog = DateRangeDialog(self)
        if dialog.exec():
            start_date, end_date = dialog.get_dates()
            if start_date > end_date: show_error_message(self, "Invalid Date Range",
                                                         "The start date cannot be after the end date."); return
            start_date_str = start_date.toString(Qt.DateFormat.ISODate);
            end_date_str = end_date.toString(Qt.DateFormat.ISODate)
            self.set_controls_enabled(False);
            self.log_audit_trail("FAILED_MONTHLY_REPORT_START",
                                 f"User initiated failed monthly report for {start_date_str} to {end_date_str}");
            self.failed_monthly_report_thread = QThread();
            self.failed_monthly_report_worker = FailedMonthlyReportWorker(engine=self.engine, start_date=start_date_str,
                                                                          end_date=end_date_str)
            self.failed_monthly_report_worker.moveToThread(self.failed_monthly_report_thread);
            self.failed_monthly_report_thread.started.connect(self.failed_monthly_report_worker.run);
            self.failed_monthly_report_worker.finished.connect(self._on_failed_monthly_report_finished);
            self.failed_monthly_report_worker.error.connect(self._on_calculation_error);
            self.failed_monthly_report_worker.finished.connect(self.failed_monthly_report_thread.quit);
            self.failed_monthly_report_thread.finished.connect(self._reset_failed_monthly_report_thread_state);
            self.failed_monthly_report_thread.start()

    def _on_failed_monthly_report_finished(self, df: pd.DataFrame):
        self.set_controls_enabled(True)
        if df.empty: QMessageBox.information(self, "Report Complete",
                                             "No failed transactions found for the selected date range."); return

        def get_remark(quantity):
            return "Net consumption from prior inventory" if quantity < 0 else "Net increase in inventory"

        df['remarks'] = df['net_quantity'].apply(get_remark)
        default_filename = f"Failed_Monthly_Transaction_Report_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Failed Monthly Report", default_filename,
                                                  "Excel Files (*.xlsx)")
        if filepath:
            try:
                export_df = df.rename(
                    columns={'product_code': 'PROD_CODE', 'lot_number': 'LOTNUMBER', 'net_quantity': 'QTY',
                             'location': 'LOCATION', 'bag_box_number': 'BAG_NUMBER', 'remarks': 'REMARKS'});
                final_cols = ['PROD_CODE', 'LOTNUMBER', 'QTY', 'LOCATION', 'BAG_NUMBER', 'REMARKS'];
                export_df = export_df[final_cols]
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='Failed Monthly Transactions');
                    self._format_excel_sheet(writer.sheets['Failed Monthly Transactions'], export_df)
                filename = os.path.basename(filepath);
                self.log_audit_trail("FAILED_MONTHLY_REPORT_EXPORT",
                                     f"User exported failed monthly report to '{filename}'.");
                QMessageBox.information(self, "Export Successful", f"Failed Monthly Report exported to:\n{filename}")
            except Exception:
                show_error_message(self, "Export Error", "Failed to save the failed monthly report file.",
                                   traceback.format_exc())

    def _reset_failed_monthly_report_thread_state(self):
        if self.failed_monthly_report_worker: self.failed_monthly_report_worker.deleteLater(); self.failed_monthly_report_worker = None
        if self.failed_monthly_report_thread: self.failed_monthly_report_thread.deleteLater(); self.failed_monthly_report_thread = None

    def _export_and_email(self):
        if self.current_inventory_df.empty: QMessageBox.information(self, "Email Failed",
                                                                    "No data available to email."); return
        settings = QSettings("MyCompany", "FGInventoryApp");
        email_config = {"sender_email": settings.value("email/sender_email", ""),
                        "sender_password": settings.value("email/sender_password", ""),
                        "smtp_server": settings.value("email/smtp_server", ""),
                        "smtp_port": settings.value("email/smtp_port", 0, type=int),
                        "recipient_email": settings.value("email/recipient_email", "")}
        if not all([email_config["sender_email"], email_config["smtp_server"], email_config["smtp_port"],
                    email_config["recipient_email"]]) or not settings.value(
            "email/sender_password"): QMessageBox.warning(self, "Configuration Needed",
                                                          "Email settings are incomplete. Please use the 'Settings' button to configure them first."); self._open_settings_dialog(); return
        export_df = self.current_inventory_df.rename(
            columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                     'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
        final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
        df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
        df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy();
        df_to_email = {}
        if not df_mb.empty: df_to_email['FG FAILED MB'] = df_mb
        if not df_dc.empty: df_to_email['FG FAILED DC'] = df_dc
        if not df_to_email: QMessageBox.warning(self, "Email Failed", "No data to include in the email."); return
        filename = f"FG_Inventory_Report_FAILED_{datetime.now():%Y-%m-%d}.xlsx";
        attachments = {filename: df_to_email};
        as_of_date_str = self.date_picker.date().toString("yyyy-MM-dd");
        email_config['subject'] = f"FG FAILED Inventory Report as of {as_of_date_str}";
        self.set_controls_enabled(False);
        self.email_thread = QThread();
        self.email_worker = EmailWorker(attachments=attachments, email_config=email_config,
                                        log_func=self.log_audit_trail, log_action="EMAIL_FAILED_INVENTORY",
                                        recipient=email_config['recipient_email'])
        self.email_worker.moveToThread(self.email_thread);
        self.email_thread.started.connect(self.email_worker.run);
        self.email_worker.finished.connect(self.email_thread.quit);
        self.email_worker.finished.connect(self.email_worker.deleteLater);
        self.email_thread.finished.connect(self.email_thread.deleteLater);
        self.email_worker.error.connect(self._on_email_error);
        self.email_worker.success.connect(self._on_email_success);
        self.email_thread.start()

    def _prepare_report_sheets(self, df: pd.DataFrame, report_status: str) -> dict:
        if df.empty: return {}
        status_upper = report_status.upper();
        export_df = df.rename(
            columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                     'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
        final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION']
        if 'fg_type' not in export_df.columns: export_df['fg_type'] = export_df['PRODUCT_CODE'].apply(
            lambda x: 'DC' if isinstance(x, str) and '-' in x else 'MB')
        df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
        df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy();
        sheet_dict = {}
        if not df_mb.empty: sheet_dict[f'FG MB {status_upper}'] = df_mb
        if not df_dc.empty: sheet_dict[f'FG DC {status_upper}'] = df_dc
        return sheet_dict

    def _export_and_email_combined_report(self):
        if not self.good_inventory_page: QMessageBox.warning(self, "Error",
                                                             "The passed inventory page is not linked. Cannot create a combined report."); return
        failed_df = self.current_inventory_df;
        passed_df = self.good_inventory_page.current_inventory_df
        if passed_df.empty and failed_df.empty: QMessageBox.information(self, "Email Failed",
                                                                        "No data available in either report to email."); return
        settings = QSettings("MyCompany", "FGInventoryApp");
        email_config = {"sender_email": settings.value("email/sender_email", ""),
                        "sender_password": settings.value("email/sender_password", ""),
                        "smtp_server": settings.value("email/smtp_server", ""),
                        "smtp_port": settings.value("email/smtp_port", 0, type=int),
                        "recipient_email": settings.value("email/recipient_email", "")}
        if not all([email_config["sender_email"], email_config["smtp_server"], email_config["smtp_port"],
                    email_config["recipient_email"]]) or not settings.value(
            "email/sender_password"): QMessageBox.warning(self, "Configuration Needed",
                                                          "Email settings are incomplete. Please use the 'Settings' button to configure them first."); self._open_settings_dialog(); return
        attachments = {};
        passed_sheets = self._prepare_report_sheets(passed_df, "PASSED")
        if passed_sheets: filename = f"FG_Inventory_Report_PASSED_{datetime.now():%Y-%m-%d}.xlsx"; attachments[
            filename] = passed_sheets
        failed_sheets = self._prepare_report_sheets(failed_df, "FAILED")
        if failed_sheets: filename = f"FG_Inventory_Report_FAILED_{datetime.now():%Y-%m-%d}.xlsx"; attachments[
            filename] = failed_sheets
        if not attachments: QMessageBox.warning(self, "Email Failed",
                                                "No data to include in the email after filtering."); return
        as_of_date_str = self.date_picker.date().toString("yyyy-MM-dd");
        email_config['subject'] = f"COMBINED Inventory Report (Passed & Failed) as of {as_of_date_str}";
        self.set_controls_enabled(False)
        self.email_thread = QThread();
        self.email_worker = EmailWorker(attachments=attachments, email_config=email_config,
                                        log_func=self.log_audit_trail, log_action="EMAIL_COMBINED_INVENTORY",
                                        recipient=email_config['recipient_email'])
        self.email_worker.moveToThread(self.email_thread);
        self.email_thread.started.connect(self.email_worker.run);
        self.email_worker.finished.connect(self.email_thread.quit);
        self.email_worker.finished.connect(self.email_worker.deleteLater);
        self.email_thread.finished.connect(self.email_thread.deleteLater);
        self.email_worker.error.connect(self._on_email_error);
        self.email_worker.success.connect(self._on_email_success);
        self.email_thread.start()

    def closeEvent(self, event):
        if self.inventory_thread and self.inventory_thread.isRunning(): self.inventory_thread.quit(); self.inventory_thread.wait(
            2000)
        if self.email_thread and self.email_thread.isRunning(): self.email_thread.quit(); self.email_thread.wait(2000)
        if self.failed_monthly_report_thread and self.failed_monthly_report_thread.isRunning(): self.failed_monthly_report_thread.quit(); self.failed_monthly_report_thread.wait(
            2000)
        event.accept()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # --- Database Connection ---
    DB_USER = "postgres";
    DB_PASS = "mbpi";
    DB_HOST = "192.168.1.13";
    DB_PORT = "5432";
    DB_NAME = "dbfg"
    POSTGRES_URL = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    try:
        real_db_engine = create_engine(POSTGRES_URL)
        with real_db_engine.connect() as conn:
            print("Successfully connected to the database.")
    except Exception as e:
        QMessageBox.critical(None, "Database Error", f"Could not connect to the database.\n\nError: {e}"); sys.exit(1)


    # --- Mocks for Standalone Run ---
    def production_log_audit_trail(action, details):
        print(f"[AUDIT LOG] User: ProductionUser, Action: {action}, Details: {details}")


    class DummyGoodInventoryPage(QObject):
        current_inventory_df = pd.DataFrame(
            {'product_code': ['PROD-A'], 'lot_number': ['LOT001'], 'current_balance': [100.5], 'fg_type': ['MB']})

        def set_controls_enabled(self, enabled: bool): print(f"DummyGoodInventoryPage controls enabled: {enabled}")


    # --- Main Application Window ---
    good_inv_page = DummyGoodInventoryPage()
    failed_inventory_page = FailedInventoryReportPage(engine=real_db_engine, username="ProductionUser",
                                                      log_audit_trail_func=production_log_audit_trail,
                                                      good_inventory_page=good_inv_page)
    main_window = QMainWindow();
    main_window.setWindowTitle("FG Inventory Report (Failed)");
    main_window.setCentralWidget(failed_inventory_page);
    main_window.resize(1200, 800);
    main_window.show()
    sys.exit(app.exec())