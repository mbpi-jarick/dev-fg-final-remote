import pandas as pd
import os
import sys
import traceback
from datetime import datetime
import qtawesome as fa

# --- Email Imports ---
from email.message import EmailMessage
from io import BytesIO
import smtplib
import ssl

# --- PyQt6 Imports ---
from PyQt6.QtCore import Qt, QObject, pyqtSignal, QThread, QDate, QSize, QSettings, QPoint
from PyQt6.QtGui import QColor, QAction
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QTableWidget,
                             QTableWidgetItem, QHeaderView, QMessageBox, QHBoxLayout,
                             QLabel, QPushButton, QLineEdit, QAbstractItemView,
                             QDateEdit, QGroupBox, QFileDialog, QTabWidget,
                             QGridLayout, QProgressBar, QComboBox, QDialog,
                             QFormLayout, QDialogButtonBox, QSpinBox, QMenu)

# --- SQLAlchemy and OpenPyXL Imports ---
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- UI CONSTANTS ---
PRIMARY_ACCENT_COLOR = "#007bff"
PRIMARY_ACCENT_HOVER = "#e9f0ff"
NEUTRAL_COLOR = "#6c757d"
LIGHT_TEXT_COLOR = "#333333"
BACKGROUND_CONTENT_COLOR = "#f4f7fc"
INPUT_BACKGROUND_COLOR = "#ffffff"
GROUP_BOX_HEADER_COLOR = "#f4f7fc"
TABLE_SELECTION_COLOR = "#3a506b"
ERROR_COLOR = "#dc3545"
HEADER_AND_ICON_COLOR = "#3a506b"
TABLE_HEADER_TEXT_COLOR = "#4f4f4f"


# --- Helper Widgets and Functions ---
class UpperCaseLineEdit(QLineEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.textEdited.connect(self._to_upper)

    def _to_upper(self, text: str):
        if text and text != self.text().upper():
            self.blockSignals(True)
            self.setText(text.upper())
            self.blockSignals(False)


class InventoryWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, product_filter: str, lot_filter: str, as_of_date: str, fg_type_filter: str):
        super().__init__()
        self.engine = engine
        self.product_filter = product_filter
        self.lot_filter = lot_filter
        self.as_of_date = as_of_date
        self.fg_type_filter = fg_type_filter

    def run(self):
        try:
            params = {'as_of_date': self.as_of_date}
            date_filter_clause = "AND t.transaction_date <= :as_of_date"
            product_filter_clause = ""
            if self.product_filter:
                product_filter_clause = "AND s.product_code ILIKE :product_search"
                params['product_search'] = f"%{self.product_filter}%"
            lot_filter_clause = ""
            if self.lot_filter:
                lot_filter_clause = "AND s.lot_number ILIKE :lot_search"
                params['lot_search'] = f"%{self.lot_filter}%"
            fg_type_clause = ""
            if self.fg_type_filter in ["MB", "DC"]:
                fg_type_clause = "AND s.fg_type = :fg_type"
                params['fg_type'] = self.fg_type_filter
            query_str = f"""
                WITH all_tx AS (
                    SELECT UPPER(TRIM(product_code)) AS product_code, UPPER(TRIM(lot_number)) AS lot_number, COALESCE(UPPER(TRIM(fg_type)), CASE WHEN product_code LIKE '%-%' THEN 'DC' ELSE 'MB' END) AS fg_type, COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in, 0.0 AS quantity_out, location
                    FROM beginv_sheet1 WHERE product_code IS NOT NULL AND TRIM(product_code) <> '' AND lot_number IS NOT NULL AND TRIM(lot_number) <> ''
                    UNION ALL
                    SELECT UPPER(TRIM(t.product_code)) AS product_code, UPPER(TRIM(t.lot_number)) AS lot_number, CASE WHEN t.product_code LIKE '%-%' THEN 'DC' ELSE 'MB' END AS fg_type, COALESCE(CAST(t.quantity_in AS NUMERIC), 0) AS quantity_in, COALESCE(CAST(t.quantity_out AS NUMERIC), 0) AS quantity_out, t.warehouse as location
                    FROM transactions t WHERE t.product_code IS NOT NULL AND TRIM(t.product_code) <> '' AND t.lot_number IS NOT NULL AND TRIM(t.lot_number) <> '' {date_filter_clause}
                ),
                lot_details_source AS (
                    ( SELECT DISTINCT ON (t.lot_number) t.lot_number, t.warehouse AS location, COALESCE(fg_b.bag_no, qcf_b.bag_no, '') AS bag_box_number, 0 as priority FROM transactions t
                        LEFT JOIN (SELECT DISTINCT ON (system_ref_no) system_ref_no, bag_no FROM fg_endorsements_primary ORDER BY system_ref_no, bag_no) fg_b ON t.source_ref_no = fg_b.system_ref_no
                        LEFT JOIN (SELECT DISTINCT ON (system_ref_no) system_ref_no, bag_no FROM qcf_endorsements_primary ORDER BY system_ref_no, bag_no) qcf_b ON t.source_ref_no = qcf_b.system_ref_no
                        WHERE t.transaction_date <= :as_of_date ORDER BY t.lot_number, t.transaction_date DESC )
                    UNION ALL
                    ( SELECT DISTINCT ON (b.lot_number) b.lot_number, b.location, COALESCE(b.bag_number, b.box_number, '') as bag_box_number, 1 as priority FROM beginv_sheet1 b
                        WHERE b.lot_number IS NOT NULL AND TRIM(b.lot_number) <> '' ORDER BY b.lot_number )
                ),
                lot_details AS (SELECT DISTINCT ON (lot_number) lot_number, location, bag_box_number FROM lot_details_source ORDER BY lot_number, priority ASC),
                lot_summary AS (
                    SELECT lot_number, MAX(product_code) AS product_code, MAX(fg_type) AS fg_type, COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0) AS current_balance
                    FROM all_tx GROUP BY lot_number HAVING (COALESCE(SUM(quantity_in), 0) - COALESCE(SUM(quantity_out), 0)) != 0
                )
                SELECT s.product_code, s.lot_number, s.current_balance, d.location, d.bag_box_number, s.fg_type
                FROM lot_summary s LEFT JOIN lot_details d ON s.lot_number = d.lot_number
                WHERE 1=1 {product_filter_clause} {lot_filter_clause} {fg_type_clause} ORDER BY s.product_code, s.lot_number;
            """
            with self.engine.connect() as conn:
                results = conn.execute(text(query_str), params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame(
                columns=['product_code', 'lot_number', 'current_balance', 'location', 'bag_box_number', 'fg_type'])
            if not df.empty:
                df['current_balance'] = pd.to_numeric(df['current_balance'])
            self.finished.emit(df)
        except Exception:
            self.error.emit("Database query failed.", traceback.format_exc())


class TransactionHistoryWorker(QObject):
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, lot_number: str, as_of_date: str):
        super().__init__()
        self.engine = engine
        self.lot_number = lot_number
        self.as_of_date = as_of_date

    def run(self):
        try:
            query = text("""
                SELECT '1900-01-01'::date AS transaction_date, 'BEGINV' AS transaction_type, COALESCE(CAST(qty AS NUMERIC), 0) AS quantity_in, 0.0 AS quantity_out, location, 'N/A' as source_ref_no
                FROM beginv_sheet1 WHERE UPPER(TRIM(lot_number)) = :lot_number
                UNION ALL
                SELECT t.transaction_date, t.transaction_type, COALESCE(CAST(t.quantity_in AS NUMERIC), 0) AS quantity_in, COALESCE(CAST(t.quantity_out AS NUMERIC), 0) AS quantity_out, t.warehouse AS location, t.source_ref_no
                FROM transactions t WHERE UPPER(TRIM(t.lot_number)) = :lot_number AND t.transaction_date <= :as_of_date
                ORDER BY transaction_date;
            """)
            params = {'lot_number': self.lot_number, 'as_of_date': self.as_of_date}
            with self.engine.connect() as conn:
                results = conn.execute(query, params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame()
            self.finished.emit(df)
        except Exception:
            self.error.emit("Lot History Query Failed", traceback.format_exc())


class MonthlyReportWorker(QObject):
    """
    MODIFIED: Calculates the Net Endorsement (Passed - Failed) aggregated by unique product code.
    """
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, start_date: str, end_date: str):
        super().__init__()
        self.engine = engine
        self.start_date = start_date
        self.end_date = end_date

    def run(self):
        try:
            params = {'start_date': self.start_date, 'end_date': self.end_date}
            query_str = f"""
                WITH all_endorsements AS (
                    -- FG Endorsements (Passed - Treated as Positive Quantity)
                    SELECT
                        UPPER(TRIM(product_code)) as product_code,
                        COALESCE(CAST("QTY" AS NUMERIC), 0) as net_qty, -- Attempting uppercase QTY
                        remarks
                    FROM fg_endorsements_primary
                    WHERE endorsement_date BETWEEN :start_date AND :end_date
                    UNION ALL
                    -- QCF Endorsements (Failed - Treated as Negative Quantity)
                    SELECT
                        UPPER(TRIM(product_code)) as product_code,
                        -1 * COALESCE(CAST("QTY" AS NUMERIC), 0) as net_qty, -- Attempting uppercase QTY
                        remarks
                    FROM qcf_endorsements_primary
                    WHERE endorsement_date BETWEEN :start_date AND :end_date
                )
                SELECT
                    product_code,
                    SUM(net_qty) as net_quantity,
                    -- Aggregate all unique non-empty remarks from both FG and QCF endorsements
                    STRING_AGG(DISTINCT NULLIF(TRIM(remarks), ''), '; ' ORDER BY product_code) AS remarks
                FROM all_endorsements
                GROUP BY product_code
                HAVING SUM(net_qty) != 0 -- Only show products that had a net movement
                ORDER BY product_code;
            """
            with self.engine.connect() as conn:
                results = conn.execute(text(query_str), params).mappings().all()

            df = pd.DataFrame(results) if results else pd.DataFrame(
                columns=['product_code', 'net_quantity', 'remarks'])

            if not df.empty:
                df['net_quantity'] = pd.to_numeric(df['net_quantity'])

            self.finished.emit(df)
        except Exception:
            self.error.emit("Monthly Report query failed.", traceback.format_exc())


class EndorsementSummaryWorker(QObject):
    """
    Calculates the net endorsement quantity (Passed - Failed) for each product
    within a date range and aggregates unique remarks from both sources.
    """
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str, str)

    def __init__(self, engine: Engine, start_date: str, end_date: str):
        super().__init__()
        self.engine = engine
        self.start_date = start_date
        self.end_date = end_date

    def run(self):
        try:
            query = text("""
                WITH all_endorsements AS (
                    -- FG Endorsements (Passed - Treated as Positive Quantity)
                    SELECT
                        UPPER(TRIM(product_code)) as product_code,
                        COALESCE(CAST("QTY" AS NUMERIC), 0) as net_qty, -- Attempting uppercase QTY
                        remarks
                    FROM fg_endorsements_primary
                    WHERE endorsement_date BETWEEN :start_date AND :end_date
                    UNION ALL
                    -- QCF Endorsements (Failed - Treated as Negative Quantity)
                    SELECT
                        UPPER(TRIM(product_code)) as product_code,
                        -1 * COALESCE(CAST("QTY" AS NUMERIC), 0) as net_qty, -- Attempting uppercase QTY
                        remarks
                    FROM qcf_endorsements_primary
                    WHERE endorsement_date BETWEEN :start_date AND :end_date
                )
                SELECT
                    product_code,
                    SUM(net_qty) as total_qty,
                    -- Aggregate all unique non-empty remarks
                    STRING_AGG(DISTINCT NULLIF(TRIM(remarks), ''), '; ' ORDER BY product_code) as aggregated_remarks
                FROM all_endorsements
                GROUP BY product_code
                HAVING SUM(net_qty) != 0 
                ORDER BY product_code;
            """)
            params = {'start_date': self.start_date, 'end_date': self.end_date}
            with self.engine.connect() as conn:
                results = conn.execute(query, params).mappings().all()
            df = pd.DataFrame(results) if results else pd.DataFrame()

            if not df.empty:
                df['total_qty'] = pd.to_numeric(df['total_qty'])

            self.finished.emit(df)
        except Exception:
            self.error.emit("Endorsement Summary Failed", traceback.format_exc())


class TransactionHistoryDialog(QDialog):
    def __init__(self, engine: Engine, lot_number: str, as_of_date: str, parent=None):
        super().__init__(parent)
        self.engine = engine
        self.lot_number = lot_number
        self.as_of_date = as_of_date
        self.worker_thread = None
        self.worker = None
        self.setWindowTitle(f"Transaction History: {self.lot_number}")
        self.setMinimumSize(800, 500)
        self.setStyleSheet(parent.styleSheet() if parent else "")
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"<b>History for Lot:</b> {self.lot_number}", objectName="PageHeader"))
        self.history_table = QTableWidget()
        self.setup_table()
        layout.addWidget(self.history_table, 1)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        self.fetch_transactions()

    def setup_table(self):
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels(
            ["Date", "Type", "Ref No.", "Qty In", "Qty Out", "Running Balance"])
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

    def fetch_transactions(self):
        self._show_loading_state("Fetching transaction history...")
        self.worker_thread = QThread()
        self.worker = TransactionHistoryWorker(self.engine, self.lot_number, self.as_of_date)
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(self._display_history)
        self.worker.error.connect(self._on_fetch_error)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker_thread.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker_thread.start()

    def _display_history(self, df: pd.DataFrame):
        self.history_table.setRowCount(0)
        if df.empty: self._show_loading_state("No transactions found for this lot."); return
        df['running_balance'] = (df['quantity_in'].astype(float) - df['quantity_out'].astype(float)).cumsum()
        self.history_table.setRowCount(len(df))
        for i, row in df.iterrows():
            date_str = row['transaction_date'].strftime('%Y-%m-%d') if pd.notna(row['transaction_date']) and row[
                'transaction_type'] != 'BEGINV' else 'Beginning'
            self.history_table.setItem(i, 0, QTableWidgetItem(date_str))
            self.history_table.setItem(i, 1, QTableWidgetItem(str(row.get('transaction_type', ''))))
            self.history_table.setItem(i, 2, QTableWidgetItem(str(row.get('source_ref_no', ''))))
            for col_idx, col_name in enumerate(['quantity_in', 'quantity_out', 'running_balance'], 3):
                item = QTableWidgetItem(f"{row.get(col_name, 0.0):,.2f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.history_table.setItem(i, col_idx, item)

    def _on_fetch_error(self, title, message):
        show_error_message(self, title, "Could not fetch lot history.", message)
        self._show_loading_state("Error fetching data.")

    def _show_loading_state(self, message: str):
        self.history_table.setRowCount(1)
        self.history_table.setSpan(0, 0, 1, self.history_table.columnCount())
        loading_item = QTableWidgetItem(message)
        loading_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.history_table.setItem(0, 0, loading_item)


def show_error_message(parent, title, message, detailed_text=""):
    msg_box = QMessageBox(parent)
    msg_box.setIcon(QMessageBox.Icon.Critical)
    msg_box.setWindowTitle(title)
    msg_box.setText(f"<b>{message}</b>")
    if detailed_text: msg_box.setDetailedText(detailed_text)
    msg_box.exec()


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Email Settings")
        self.setMinimumWidth(450)
        self.settings = QSettings("MyCompany", "FGInventoryApp")
        main_layout = QVBoxLayout(self)
        sender_group = QGroupBox("Sender Details (SMTP)")
        sender_layout = QFormLayout(sender_group)
        self.sender_email_input = QLineEdit()
        self.sender_password_input = QLineEdit()
        self.sender_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.smtp_server_input = QLineEdit()
        self.smtp_port_input = QSpinBox()
        self.smtp_port_input.setRange(1, 65535)
        sender_layout.addRow("Sender Email:", self.sender_email_input)
        sender_layout.addRow("Password:", self.sender_password_input)
        sender_layout.addRow("SMTP Server:", self.smtp_server_input)
        sender_layout.addRow("SMTP Port:", self.smtp_port_input)
        main_layout.addWidget(sender_group)
        recipient_group = QGroupBox("Default Recipient")
        recipient_layout = QFormLayout(recipient_group)
        self.recipient_email_input = QLineEdit()
        self.recipient_email_input.setPlaceholderText("e.g., user1@example.com, user2@example.com")
        recipient_layout.addRow("Recipient Email(s):", self.recipient_email_input)
        main_layout.addWidget(recipient_group)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        main_layout.addWidget(button_box)
        self.load_settings()

    def load_settings(self):
        self.sender_email_input.setText(self.settings.value("email/sender_email", ""))
        if self.settings.value("email/sender_password"): self.sender_password_input.setPlaceholderText(
            "Password is saved. Enter new to change.")
        self.smtp_server_input.setText(self.settings.value("email/smtp_server", ""))
        self.smtp_port_input.setValue(int(self.settings.value("email/smtp_port", 587)))
        self.recipient_email_input.setText(self.settings.value("email/recipient_email", ""))

    def accept(self):
        self.settings.setValue("email/sender_email", self.sender_email_input.text().strip())
        if self.sender_password_input.text(): self.settings.setValue("email/sender_password",
                                                                     self.sender_password_input.text())
        self.settings.setValue("email/smtp_server", self.smtp_server_input.text().strip())
        self.settings.setValue("email/smtp_port", self.smtp_port_input.value())
        self.settings.setValue("email/recipient_email", self.recipient_email_input.text().strip())
        QMessageBox.information(self, "Success", "Settings have been saved.")
        super().accept()


class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Report Date Range")
        self.setMinimumWidth(350)
        layout = QFormLayout(self)
        self.start_date_edit = QDateEdit(calendarPopup=True, displayFormat="yyyy-MM-dd")
        self.end_date_edit = QDateEdit(calendarPopup=True, displayFormat="yyyy-MM-dd")
        today = QDate.currentDate()
        first_day_of_month = QDate(today.year(), today.month(), 1)
        self.start_date_edit.setDate(first_day_of_month)
        self.end_date_edit.setDate(today)
        layout.addRow("Start Date:", self.start_date_edit)
        layout.addRow("End Date:", self.end_date_edit)
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_dates(self):
        return self.start_date_edit.date(), self.end_date_edit.date()


def send_email_with_excel(attachments, email_config):
    sender_email = email_config.get('sender_email');
    sender_password = email_config.get('sender_password');
    smtp_server = email_config.get('smtp_server');
    smtp_port = email_config.get('smtp_port');
    recipient_email = email_config.get('recipient_email');
    subject = email_config.get('subject')
    if not all([sender_email, sender_password, smtp_server, smtp_port, recipient_email]): raise ValueError(
        "Incomplete email settings. Please configure them in the Settings menu.")
    recipients = [email.strip() for email in recipient_email.split(',')];
    msg = EmailMessage();
    msg['Subject'] = subject;
    msg['From'] = sender_email;
    msg['To'] = ', '.join(recipients)
    attached_files = list(attachments.keys());
    body = f'Please find the attached Excel report(s): {", ".join(attached_files)}.\n\nThis is an automated message.';
    msg.set_content(body)
    for filename, df_dict in attachments.items():
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for sheet_name, df in df_dict.items(): df.to_excel(writer, sheet_name=sheet_name, index=False)
        excel_buffer.seek(0)
        msg.add_attachment(excel_buffer.read(), maintype='application',
                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)
    context = ssl.create_default_context();
    smtp_port_int = int(smtp_port)
    if smtp_port_int == 465:
        with smtplib.SMTP_SSL(smtp_server, smtp_port_int, context=context) as server:
            server.login(sender_email, sender_password);
            server.send_message(msg, from_addr=sender_email,
                                to_addrs=recipients)
    else:
        with smtplib.SMTP(smtp_server, smtp_port_int) as server:
            server.starttls(context=context);
            server.login(sender_email, sender_password);
            server.send_message(msg,
                                from_addr=sender_email,
                                to_addrs=recipients)


class DashboardWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.data_df = pd.DataFrame()
        self.init_ui()
        self.setStyleSheet(self._get_dashboard_styles())

    def init_ui(self):
        main_layout = QVBoxLayout(self);
        main_layout.setContentsMargins(10, 10, 10, 10);
        main_layout.setSpacing(15);
        grid_layout = QGridLayout();
        grid_layout.setRowStretch(2, 1);
        grid_layout.setColumnStretch(0, 1);
        grid_layout.setColumnStretch(1, 1);
        summary_group = QGroupBox("Overall Summary Metrics");
        summary_group.setObjectName("SummaryGroup");
        summary_layout = QHBoxLayout(summary_group);
        self.total_lots_label = self._create_summary_box("Total Lots:", "0");
        self.total_products_label = self._create_summary_box("Unique Products:", "0");
        self.overall_balance_label = self._create_summary_box("Overall Balance (kg):", "0.00");
        summary_layout.addWidget(self.total_lots_label);
        summary_layout.addWidget(self.total_products_label);
        summary_layout.addWidget(self.overall_balance_label);
        grid_layout.addWidget(summary_group, 0, 0, 1, 2);
        self.contribution_table = self._create_contribution_table("product");
        contribution_group = QGroupBox("Top 10 Product Contribution (by Mass)");
        contribution_group.setObjectName("ContributionGroup");
        vbox_contribution = QVBoxLayout(contribution_group);
        vbox_contribution.addWidget(self.contribution_table);
        grid_layout.addWidget(contribution_group, 1, 0);
        self.location_table = self._create_contribution_table("location");
        location_group = QGroupBox("Top 5 Locations by Mass");
        location_group.setObjectName("LocationGroup");
        vbox_location = QVBoxLayout(location_group);
        vbox_location.addWidget(self.location_table);
        grid_layout.addWidget(location_group, 1, 1);
        self.lot_stats_group = self._create_lot_statistics_group();
        grid_layout.addWidget(self.lot_stats_group, 2, 0);
        main_layout.addLayout(grid_layout)

    def _get_dashboard_styles(self) -> str:
        return f"QGroupBox#SummaryGroup, QGroupBox#ContributionGroup, QGroupBox#LotStatsGroup, QGroupBox#LocationGroup{{border: 1px solid #e0e5eb; border-radius: 8px;margin-top: 12px; background-color: {INPUT_BACKGROUND_COLOR};}} QGroupBox::title{{subcontrol-origin: margin; subcontrol-position: top left;padding: 2px 10px; background-color: {GROUP_BOX_HEADER_COLOR};border: 1px solid #e0e5eb; border-bottom: none;border-top-left-radius: 8px; border-top-right-radius: 8px;font-weight: bold; color: {HEADER_AND_ICON_COLOR};}} QLabel#TitleLabel{{ font-size: 10pt; color: {HEADER_AND_ICON_COLOR}; background-color: transparent; }} QLabel#ValueLabel{{ font-size: 16pt; font-weight: bold; color: {HEADER_AND_ICON_COLOR}; background-color: transparent; }} QTableWidget{{ border: 1px solid #e0e5eb; background-color: {INPUT_BACKGROUND_COLOR}; gridline-color: #f0f3f8;}} QTableWidget::item:selected{{ background-color: {TABLE_SELECTION_COLOR}; color: white;}} QHeaderView::section{{ background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR};}} QProgressBar::chunk{{ background-color: {PRIMARY_ACCENT_COLOR}; border-radius: 4px; }}"

    def _create_summary_box(self, title: str, initial_value: str) -> QWidget:
        widget = QWidget();
        layout = QVBoxLayout(widget);
        layout.setContentsMargins(15, 10, 15, 10);
        title_label = QLabel(title, objectName="TitleLabel");
        value_label = QLabel(initial_value, objectName="ValueLabel", alignment=Qt.AlignmentFlag.AlignCenter);
        layout.addWidget(title_label, alignment=Qt.AlignmentFlag.AlignCenter);
        layout.addWidget(value_label, alignment=Qt.AlignmentFlag.AlignCenter);
        widget.setStyleSheet(
            f"background-color: {INPUT_BACKGROUND_COLOR}; border: 1px solid #d1d9e6; border-radius: 8px;");
        return widget

    def _create_contribution_table(self, table_type: str) -> QTableWidget:
        table = QTableWidget(columnCount=4);
        header_label = "Product" if table_type == "product" else "Location";
        table.setHorizontalHeaderLabels([header_label, "Balance (kg)", "% Share", ""]);
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers);
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows);
        table.verticalHeader().setVisible(False);
        table.setShowGrid(True);
        table.setAlternatingRowColors(True);
        header = table.horizontalHeader();
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch);
        return table

    def _create_lot_statistics_group(self) -> QGroupBox:
        group = QGroupBox("Lot Size Statistics", objectName="LotStatsGroup");
        layout = QGridLayout(group);
        self.lot_stats = {'max_lot': QLabel("N/A"), 'min_lot': QLabel("N/A"), 'avg_lot': QLabel("N/A"),
                          'median_lot': QLabel("N/A")};
        label_style = f"font-size:10pt;color:{HEADER_AND_ICON_COLOR};background-color:transparent;";
        value_style = f"font-weight:bold;color:{HEADER_AND_ICON_COLOR};background-color:transparent;";
        layout.addWidget(QLabel("Largest Lot (kg):", styleSheet=label_style), 0, 0);
        self.lot_stats['max_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['max_lot'], 0, 1, alignment=Qt.AlignmentFlag.AlignRight);
        layout.addWidget(QLabel("Smallest Lot (kg):", styleSheet=label_style), 1, 0);
        self.lot_stats['min_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['min_lot'], 1, 1, alignment=Qt.AlignmentFlag.AlignRight);
        layout.addWidget(QLabel("Average Lot (kg):", styleSheet=label_style), 2, 0);
        self.lot_stats['avg_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['avg_lot'], 2, 1, alignment=Qt.AlignmentFlag.AlignRight);
        layout.addWidget(QLabel("Median Lot (kg):", styleSheet=label_style), 3, 0);
        self.lot_stats['median_lot'].setStyleSheet(value_style);
        layout.addWidget(self.lot_stats['median_lot'], 3, 1, alignment=Qt.AlignmentFlag.AlignRight);
        layout.setRowStretch(4, 1);
        return group

    def update_dashboard(self, df: pd.DataFrame):
        self.data_df = df;
        is_empty = df.empty;
        positive_balance_df = df[df['current_balance'] > 0] if not is_empty else pd.DataFrame();
        overall_balance = df['current_balance'].sum() if not is_empty else 0.0
        self.total_lots_label.findChild(QLabel, "ValueLabel").setText("0" if is_empty else f"{len(df):,}");
        self.total_products_label.findChild(QLabel, "ValueLabel").setText(
            "0" if is_empty else f"{df['product_code'].nunique():,}");
        self.overall_balance_label.findChild(QLabel, "ValueLabel").setText(f"{overall_balance:,.2f}")
        self.lot_stats['max_lot'].setText(
            "N/A" if positive_balance_df.empty else f"{positive_balance_df['current_balance'].max():,.2f}");
        self.lot_stats['min_lot'].setText(
            "N/A" if positive_balance_df.empty else f"{positive_balance_df['current_balance'].min():,.2f}");
        self.lot_stats['avg_lot'].setText(
            "N/A" if positive_balance_df.empty else f"{positive_balance_df['current_balance'].mean():,.2f}");
        self.lot_stats['median_lot'].setText(
            "N/A" if positive_balance_df.empty else f"{positive_balance_df['current_balance'].median():,.2f}")
        self.contribution_table.setRowCount(0);
        self.location_table.setRowCount(0)
        if not positive_balance_df.empty:
            positive_total = positive_balance_df['current_balance'].sum();
            self._populate_summary_table(positive_balance_df, self.contribution_table, 'product_code', positive_total,
                                         10);
            self._populate_summary_table(positive_balance_df, self.location_table, 'location', positive_total, 5)

    def _populate_summary_table(self, df, table_widget, group_by_col, total_val, top_n):
        if df.empty or total_val <= 0: return
        summary = df.groupby(group_by_col)['current_balance'].sum().nlargest(top_n).reset_index();
        summary['percentage'] = (summary['current_balance'] / total_val) * 100
        table_widget.setRowCount(len(summary))
        for i, row in summary.iterrows():
            table_widget.setItem(i, 0, QTableWidgetItem(str(row[group_by_col])));
            balance_item = QTableWidgetItem(f"{row['current_balance']:,.2f}");
            balance_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter);
            table_widget.setItem(i, 1, balance_item);
            percent_item = QTableWidgetItem(f"{row['percentage']:.1f}%");
            percent_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter);
            table_widget.setItem(i, 2, percent_item);
            progress_bar = QProgressBar(maximum=100, value=int(row['percentage']), textVisible=False);
            table_widget.setCellWidget(i, 3, progress_bar)


class GoodInventoryPage(QWidget):
    def __init__(self, engine: Engine, username: str, log_audit_trail_func):
        super().__init__()
        self.engine = engine;
        self.username = username;
        self.log_audit_trail = log_audit_trail_func;
        self.failed_inventory_page = None
        self.inventory_thread: QThread | None = None;
        self.inventory_worker: InventoryWorker | None = None
        self.current_inventory_df = pd.DataFrame()
        self.monthly_report_thread: QThread | None = None;
        self.monthly_report_worker: MonthlyReportWorker | None = None
        self.endorsement_summary_thread: QThread | None = None
        self.endorsement_summary_worker: EndorsementSummaryWorker | None = None
        self.init_ui()
        self.setStyleSheet(self._get_styles())

    def refresh_page(self):
        self._start_inventory_calculation()

    def _get_styles(self) -> str:
        return f"QWidget{{background-color:{BACKGROUND_CONTENT_COLOR}; color:{LIGHT_TEXT_COLOR}; font-family:'Segoe UI',Arial,sans-serif;}} QGroupBox{{border:1px solid #e0e5eb; border-radius:8px; margin-top:12px; background-color:{INPUT_BACKGROUND_COLOR};}} QGroupBox::title{{subcontrol-origin:margin; subcontrol-position:top left; padding:2px 10px; background-color:{GROUP_BOX_HEADER_COLOR}; border:1px solid #e0e5eb; border-bottom:1px solid {INPUT_BACKGROUND_COLOR}; border-top-left-radius:8px; border-top-right-radius:8px; font-weight:bold; color:#4f4f4f;}} QGroupBox QLabel{{background-color: transparent;}} QLabel#PageHeader{{font-size:15pt; font-weight:bold; color:{HEADER_AND_ICON_COLOR}; background-color:transparent;}} QLineEdit, QDateEdit, QComboBox{{border:1px solid #d1d9e6; padding:8px; border-radius:5px; background-color:{INPUT_BACKGROUND_COLOR};}} QLineEdit:focus, QDateEdit:focus, QComboBox:focus{{border:1px solid {PRIMARY_ACCENT_COLOR};}} QPushButton{{border:1px solid #d1d9e6; padding:8px 15px; border-radius:6px; font-weight:bold; background-color:{INPUT_BACKGROUND_COLOR};}} QPushButton:hover{{background-color:#f0f3f8;}} QPushButton#PrimaryButton{{color:{HEADER_AND_ICON_COLOR}; border:1px solid {HEADER_AND_ICON_COLOR};}} QPushButton#PrimaryButton:hover{{background-color:#e9f0ff;}} QTabWidget::pane{{border:1px solid #e0e5eb; border-radius:8px; background-color:{INPUT_BACKGROUND_COLOR}; padding:10px; margin-top:-1px;}} QTabBar::tab{{background:#e9eff7; color:{NEUTRAL_COLOR}; padding:8px 15px; border:1px solid #e0e5eb; border-bottom:none; border-top-left-radius:6px; border-top-right-radius:6px;}} QTabBar::tab:selected{{color:{HEADER_AND_ICON_COLOR}; background:{INPUT_BACKGROUND_COLOR}; border-bottom-color:{INPUT_BACKGROUND_COLOR}; font-weight:bold;}} QTableWidget{{border:1px solid #e0e5eb; background-color:{INPUT_BACKGROUND_COLOR}; selection-behavior:SelectRows; gridline-color: #f0f3f8;}} QTableWidget::item:hover{{background-color: transparent;}} QTableWidget::item:selected{{background-color:{TABLE_SELECTION_COLOR}; color:white;}} QHeaderView::section{{background-color: #f4f7fc; padding: 5px; border: none; font-weight: bold; color: {TABLE_HEADER_TEXT_COLOR};}} QLabel#TotalBalanceLabel{{background-color:{INPUT_BACKGROUND_COLOR}; color:{PRIMARY_ACCENT_COLOR}; padding:10px; border-radius:6px; border:1px solid #e0e5eb; font-weight:bold;}}"

    def init_ui(self):
        main_layout = QVBoxLayout(self);
        header_widget = QWidget();
        header_layout = QHBoxLayout(header_widget);
        header_layout.setContentsMargins(0, 0, 0, 0);
        icon_label = QLabel();
        icon_label.setPixmap(fa.icon('fa5s.boxes', color=HEADER_AND_ICON_COLOR).pixmap(QSize(28, 28)));
        header_layout.addWidget(icon_label);
        header_layout.addWidget(QLabel("FG Inventory Computation and Export", objectName="PageHeader"));
        header_layout.addStretch();
        main_layout.addWidget(header_widget);
        instruction_label = QLabel(
            "This page calculates and displays the real-time inventory balance based on beginning inventory and all subsequent transactions up to the selected 'As Of Date'.");
        instruction_label.setStyleSheet("font-style: italic; color: #555; background: transparent;");
        main_layout.addWidget(instruction_label);
        main_layout.addWidget(self._create_filter_controls());
        self.tab_widget = QTabWidget();
        self.tab_widget.setIconSize(QSize(16, 16));
        self.inventory_tab = QWidget();
        inventory_layout = QVBoxLayout(self.inventory_tab);
        self.inventory_table = self._create_inventory_table();
        self.inventory_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu);
        self.inventory_table.customContextMenuRequested.connect(self.show_table_context_menu);
        self.total_balance_label = QLabel("Total Balance: 0.00 kg", objectName="TotalBalanceLabel",
                                          alignment=Qt.AlignmentFlag.AlignRight);
        inventory_layout.addWidget(self.inventory_table);
        inventory_layout.addWidget(self.total_balance_label);
        self.tab_widget.addTab(self.inventory_tab, fa.icon('fa5.list-alt', color=HEADER_AND_ICON_COLOR),
                               "Inventory Details");
        self.dashboard_widget = DashboardWidget();
        self.tab_widget.addTab(self.dashboard_widget, fa.icon('fa5s.chart-area', color=HEADER_AND_ICON_COLOR),
                               "Dashboard");
        main_layout.addWidget(self.tab_widget, 1)

    def show_table_context_menu(self, pos: QPoint):
        item = self.inventory_table.itemAt(pos);
        if not item: return
        row_index = item.row();
        lot_item = self.inventory_table.item(row_index, 1)
        if not lot_item: return
        lot_number = lot_item.text();
        context_menu = QMenu(self);
        track_action = QAction(fa.icon('fa5s.history', color=HEADER_AND_ICON_COLOR), f"Track Lot: {lot_number}", self);
        track_action.triggered.connect(lambda: self.track_lot_transactions(lot_number));
        context_menu.addAction(track_action);
        context_menu.exec(self.inventory_table.mapToGlobal(pos))

    def track_lot_transactions(self, lot_number: str):
        as_of_date_str = self.date_picker.date().toString(Qt.DateFormat.ISODate);
        dialog = TransactionHistoryDialog(self.engine, lot_number, as_of_date_str, self);
        dialog.exec()

    def _create_filter_controls(self) -> QGroupBox:
        group = QGroupBox("Filters & Actions");
        layout = QHBoxLayout(group);
        layout.addWidget(QLabel("As Of Date:"));
        self.date_picker = QDateEdit(calendarPopup=True, date=QDate.currentDate(), displayFormat="yyyy-MM-dd");
        layout.addWidget(self.date_picker);
        layout.addSpacing(10);
        layout.addWidget(QLabel("FG Type:"));
        self.fg_type_combo = QComboBox();
        self.fg_type_combo.addItems(["All", "MB", "DC"]);
        layout.addWidget(self.fg_type_combo);
        layout.addSpacing(10);
        layout.addWidget(QLabel("Product:"));
        self.product_code_input = UpperCaseLineEdit(placeholderText="Filter by Product Code");
        layout.addWidget(self.product_code_input, 1);
        layout.addSpacing(10);
        layout.addWidget(QLabel("Lot:"));
        self.lot_number_input = UpperCaseLineEdit(placeholderText="Filter by Lot Number");
        layout.addWidget(self.lot_number_input, 1);
        layout.addStretch();
        self.refresh_button = QPushButton("Refresh", objectName="PrimaryButton",
                                          icon=fa.icon('fa5s.sync-alt', color=HEADER_AND_ICON_COLOR));
        self.export_button = QPushButton("Export to Excel", objectName="PrimaryButton",
                                         icon=fa.icon('fa5s.file-excel', color=HEADER_AND_ICON_COLOR));
        self.email_button = QPushButton("Email (Passed Only)", objectName="PrimaryButton",
                                        icon=fa.icon('fa5s.paper-plane', color=HEADER_AND_ICON_COLOR));
        self.email_combined_button = QPushButton("Email Combined Report", objectName="PrimaryButton",
                                                 icon=fa.icon('fa5s.envelope-open-text', color=HEADER_AND_ICON_COLOR));
        self.monthly_report_button = QPushButton("Monthly Report (Summary)", objectName="PrimaryButton",
                                                 icon=fa.icon('fa5s.calendar-alt', color=HEADER_AND_ICON_COLOR));
        self.endorsement_report_button = QPushButton("Endorsement Summary", objectName="PrimaryButton",
                                                     icon=fa.icon('fa5s.file-signature', color=HEADER_AND_ICON_COLOR));
        self.settings_button = QPushButton("Settings", objectName="PrimaryButton",
                                           icon=fa.icon('fa5s.cog', color=HEADER_AND_ICON_COLOR));
        layout.addWidget(self.refresh_button);
        layout.addWidget(self.export_button);
        layout.addWidget(self.email_button);
        layout.addWidget(self.email_combined_button);
        layout.addWidget(self.monthly_report_button);
        layout.addWidget(self.endorsement_report_button);
        layout.addWidget(self.settings_button);
        self.refresh_button.clicked.connect(self._start_inventory_calculation);
        self.product_code_input.returnPressed.connect(self._start_inventory_calculation);
        self.lot_number_input.returnPressed.connect(self._start_inventory_calculation);
        self.date_picker.dateChanged.connect(self._start_inventory_calculation);
        self.fg_type_combo.currentIndexChanged.connect(self._start_inventory_calculation);
        self.export_button.clicked.connect(self._export_to_excel);
        self.email_button.clicked.connect(self._export_and_email);
        self.email_combined_button.clicked.connect(self._export_and_email_combined_report);
        self.monthly_report_button.clicked.connect(self._handle_monthly_report_request);
        self.endorsement_report_button.clicked.connect(self._handle_endorsement_summary_request);
        self.settings_button.clicked.connect(self._open_settings_dialog);
        return group

    def _create_inventory_table(self):
        table = QTableWidget(columnCount=5);
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers);
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows);
        table.verticalHeader().setVisible(False);
        table.setAlternatingRowColors(True);
        table.setHorizontalHeaderLabels(["PRODUCT", "LOT NUMBER", "QTY (kg)", "BAG/BOX NO.", "LOCATION"]);
        header = table.horizontalHeader();
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch);
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents);
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents);
        return table

    def _start_inventory_calculation(self):
        if self.inventory_thread and self.inventory_thread.isRunning(): return
        self.set_controls_enabled(False);
        self._show_loading_state(self.inventory_table, "Calculating inventory balance...");
        self.dashboard_widget.update_dashboard(pd.DataFrame());
        date_str = self.date_picker.date().toString(Qt.DateFormat.ISODate);
        filters = f"Date: {date_str}, Type: {self.fg_type_combo.currentText()}, Prod: '{self.product_code_input.text()}', Lot: '{self.lot_number_input.text()}'";
        self.log_audit_trail("CALCULATE_GOOD_INVENTORY", f"User calculated good inventory with filters: {filters}");
        self.inventory_thread = QThread();
        self.inventory_worker = InventoryWorker(engine=self.engine,
                                                product_filter=self.product_code_input.text().strip().upper(),
                                                lot_filter=self.lot_number_input.text().strip().upper(),
                                                as_of_date=date_str, fg_type_filter=self.fg_type_combo.currentText());
        self.inventory_worker.moveToThread(self.inventory_thread);
        self.inventory_thread.started.connect(self.inventory_worker.run);
        self.inventory_worker.finished.connect(self._on_inventory_finished);
        self.inventory_worker.error.connect(self._on_calculation_error);
        self.inventory_worker.finished.connect(self.inventory_thread.quit);
        self.inventory_worker.error.connect(self.inventory_thread.quit);
        self.inventory_thread.finished.connect(self.inventory_worker.deleteLater);
        self.inventory_thread.finished.connect(self.inventory_thread.deleteLater);
        self.inventory_thread.finished.connect(self._reset_thread_state);
        self.inventory_thread.start()

    def _reset_thread_state(self):
        self.inventory_thread = None;
        self.inventory_worker = None

    def _on_inventory_finished(self, df: pd.DataFrame):
        self.current_inventory_df = df.copy();
        self._display_inventory(df);
        self.dashboard_widget.update_dashboard(
            df);
        self.set_controls_enabled(True)

    def _on_calculation_error(self, error_message: str, detailed_traceback: str):
        show_error_message(self, "Calculation Error", error_message, detailed_traceback);
        self._show_loading_state(
            self.inventory_table, "Error during calculation.");
        self.dashboard_widget.update_dashboard(
            pd.DataFrame());
        self.set_controls_enabled(True)

    def _show_loading_state(self, table: QTableWidget, message: str):
        table.setRowCount(1);
        table.setSpan(0, 0, 1, table.columnCount());
        loading_item = QTableWidgetItem(
            message);
        loading_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter);
        table.setItem(0, 0, loading_item)

    def _display_inventory(self, df: pd.DataFrame):
        self.inventory_table.setRowCount(0)
        if df.empty: self._show_loading_state(self.inventory_table,
                                              "No inventory found."); self.total_balance_label.setText(
            "Total Balance: 0.00 kg"); return
        total_balance = df['current_balance'].sum();
        self.inventory_table.setRowCount(len(df))
        for i, row in df.iterrows():
            current_balance = row.get('current_balance', 0.0);
            qty_item = QTableWidgetItem(f"{current_balance:,.2f}");
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if current_balance < 0: qty_item.setForeground(QColor(ERROR_COLOR))
            self.inventory_table.setItem(i, 0, QTableWidgetItem(str(row.get('product_code', ''))));
            self.inventory_table.setItem(i, 1, QTableWidgetItem(str(row.get('lot_number', ''))));
            self.inventory_table.setItem(i, 2, qty_item);
            self.inventory_table.setItem(i, 3, QTableWidgetItem(str(row.get('bag_box_number', ''))));
            self.inventory_table.setItem(i, 4, QTableWidgetItem(str(row.get('location', 'N/A'))))
        is_filtered = any(
            [self.lot_number_input.text(), self.product_code_input.text(), self.fg_type_combo.currentText() != "All"]);
        prefix = f"Filtered Total ({len(df)} lots)" if is_filtered else f"Overall Total ({len(df)} lots)";
        self.total_balance_label.setText(f"{prefix}: {total_balance:,.2f} kg")

    def _format_excel_sheet(self, worksheet: Worksheet, df: pd.DataFrame):
        # Determine which column holds the quantity data for formatting
        qty_col_header = next((col for col in df.columns if 'QTY' in col.upper()), None)

        if qty_col_header and not df.empty:
            # Find the column index (1-based for openpyxl)
            qty_col_idx = df.columns.get_loc(qty_col_header) + 1
            qty_col_letter = get_column_letter(qty_col_idx)

            # Apply number format to all quantity cells (excluding header row)
            for cell in worksheet[qty_col_letter][1:]:
                cell.number_format = '#,##0.00'

        # Auto-size columns based on content length
        for i, col in enumerate(df.columns, 1):
            if not df.empty:
                column_letter = get_column_letter(i)
                # Calculate max length, minimum of 10, maximum of 50
                max_length = min(50,
                                 max(10, df[col].astype(str).map(len).max() if not df[col].empty else 10, len(col) + 2))
                worksheet.column_dimensions[column_letter].width = max_length

    def _export_to_excel(self):
        if self.current_inventory_df.empty: QMessageBox.information(self, "Export Failed", "No data to export."); return
        default_filename = f"FG_Inventory_Report_PASSED_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Report", default_filename, "Excel Files (*.xlsx)")
        if filepath:
            try:
                export_df = self.current_inventory_df.rename(
                    columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                             'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
                final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
                df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
                df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy()
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    if not df_mb.empty:
                        df_mb.to_excel(writer, index=False, sheet_name='FG MB PASSED')
                        self._format_excel_sheet(writer.sheets['FG MB PASSED'], df_mb)
                    if not df_dc.empty:
                        df_dc.to_excel(writer, index=False, sheet_name='FG DC PASSED')
                        self._format_excel_sheet(writer.sheets['FG DC PASSED'], df_dc)
                filename = os.path.basename(filepath);
                self.log_audit_trail("EXPORT_GOOD_INVENTORY", f"User exported good inventory report to '{filename}'.");
                QMessageBox.information(self, "Export Successful", f"Data exported to:\n{filename}")
            except Exception:
                show_error_message(self, "Export Error", "Failed to save file.", traceback.format_exc())

    def set_controls_enabled(self, enabled: bool):
        widgets_to_toggle = [self.lot_number_input, self.product_code_input, self.refresh_button, self.date_picker,
                             self.fg_type_combo, self.settings_button, self.email_combined_button, self.email_button,
                             self.monthly_report_button, self.endorsement_report_button]
        for widget in widgets_to_toggle: widget.setEnabled(enabled)
        is_data_present = not self.current_inventory_df.empty;
        self.export_button.setEnabled(enabled and is_data_present);
        self.email_button.setEnabled(enabled and is_data_present)

    def _handle_monthly_report_request(self):
        dialog = DateRangeDialog(self);
        if dialog.exec():
            start_date, end_date = dialog.get_dates()
            if start_date > end_date: show_error_message(self, "Invalid Date Range",
                                                         "The start date cannot be after the end date."); return
            start_date_str = start_date.toString(Qt.DateFormat.ISODate);
            end_date_str = end_date.toString(Qt.DateFormat.ISODate);

            # --- ADVICE ON PERMISSION ERROR ---
            QMessageBox.information(self, "File Save Note",
                                    "Please ensure the target file is NOT open in Excel before saving, or you will receive a 'Permission denied' error.")
            # --- END ADVICE ---

            self.set_controls_enabled(False);
            QApplication.processEvents();
            self.log_audit_trail("MONTHLY_REPORT_START",
                                 f"User initiated monthly product summary (Net Endorsement) report for {start_date_str} to {end_date_str}");
            self.monthly_report_thread = QThread();
            self.monthly_report_worker = MonthlyReportWorker(engine=self.engine, start_date=start_date_str,
                                                             end_date=end_date_str);
            self.monthly_report_worker.moveToThread(self.monthly_report_thread);
            self.monthly_report_thread.started.connect(self.monthly_report_worker.run);
            self.monthly_report_worker.finished.connect(self._on_monthly_report_finished);
            self.monthly_report_worker.error.connect(self._on_calculation_error);
            self.monthly_report_worker.finished.connect(self.monthly_report_thread.quit);
            self.monthly_report_thread.finished.connect(self._reset_monthly_report_thread_state);
            self.monthly_report_thread.start()

    def _on_monthly_report_finished(self, df: pd.DataFrame):
        self.set_controls_enabled(True)
        if df.empty:
            QMessageBox.information(self, "Report Complete",
                                    "No net endorsement movements found for the selected date range.");
            return

        default_filename = f"Monthly_Product_Summary_Report_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Monthly Product Summary", default_filename,
                                                  "Excel Files (*.xlsx)")

        if filepath:
            try:
                # Renaming columns based on new product summary structure (using Net Endorsement data)
                export_df = df.rename(
                    columns={'product_code': 'PROD_CODE',
                             'net_quantity': 'QTY (Net Endorsement Change)',
                             'remarks': 'ENDORSEMENT REMARKS'});

                # Enforce required column order
                final_cols = ['PROD_CODE', 'QTY (Net Endorsement Change)', 'ENDORSEMENT REMARKS'];
                export_df = export_df[final_cols]

                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='Monthly Endorsement Net');
                    self._format_excel_sheet(writer.sheets['Monthly Endorsement Net'], export_df)

                filename = os.path.basename(filepath);
                self.log_audit_trail("MONTHLY_SUMMARY_EXPORT",
                                     f"User exported monthly product summary (net endorsement) to '{filename}'.");
                QMessageBox.information(self, "Export Successful", f"Monthly Product Summary exported to:\n{filename}")
            except Exception:
                show_error_message(self, "Export Error",
                                   "Failed to save the monthly report file. Check if the file is open.",
                                   traceback.format_exc())

    def _reset_monthly_report_thread_state(self):
        if self.monthly_report_worker: self.monthly_report_worker.deleteLater(); self.monthly_report_worker = None
        if self.monthly_report_thread: self.monthly_report_thread.deleteLater(); self.monthly_report_thread = None

    def _open_settings_dialog(self):
        dialog = SettingsDialog(self);
        dialog.exec()

    def _export_and_email(self):
        if self.current_inventory_df.empty: QMessageBox.information(self, "Email Failed",
                                                                    "No data available to email."); return
        settings = QSettings("MyCompany", "FGInventoryApp");
        email_config = {"sender_email": settings.value("email/sender_email", ""),
                        "sender_password": settings.value("email/sender_password", ""),
                        "smtp_server": settings.value("email/smtp_server", ""),
                        "smtp_port": settings.value("email/smtp_port", 0, type=int),
                        "recipient_email": settings.value("email/recipient_email", "")}
        if not all([email_config["sender_email"], email_config["sender_password"], email_config["smtp_server"],
                    email_config["smtp_port"], email_config["recipient_email"]]): QMessageBox.warning(self,
                                                                                                      "Configuration Needed",
                                                                                                      "Email settings are incomplete. Please use the 'Settings' button to configure them first."); self._open_settings_dialog(); return
        try:
            export_df = self.current_inventory_df.rename(
                columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                         'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
            final_cols = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
            df_mb = export_df[export_df['fg_type'] == 'MB'][final_cols].copy();
            df_dc = export_df[export_df['fg_type'] == 'DC'][final_cols].copy();
            df_dict = {};
            if not df_mb.empty: df_dict['FG MB PASSED'] = df_mb
            if not df_dc.empty: df_dict['FG DC PASSED'] = df_dc
            if not df_dict: QMessageBox.warning(self, "Email Failed",
                                                "No data to include in the email after filtering."); return
            filename = f"FG_Inventory_Report_PASSED_{datetime.now():%Y-%m-%d}.xlsx";
            attachments = {filename: df_dict};
            as_of_date_str = self.date_picker.date().toString("yyyy-MM-dd");
            email_config['subject'] = f"FG Inventory Report as of {as_of_date_str}";
            self.set_controls_enabled(False);
            QApplication.processEvents();
            send_email_with_excel(attachments, email_config);
            self.log_audit_trail("EMAIL_GOOD_INVENTORY",
                                 f"User emailed good inventory report to '{email_config['recipient_email']}'.");
            QMessageBox.information(self, "Email Sent",
                                    f"The report has been successfully sent to {email_config['recipient_email']}.")
        except Exception as e:
            show_error_message(self, "Email Error", "Failed to send email. Please check your settings and connection.",
                               f"Error details:\n\n{traceback.format_exc()}")
        finally:
            self.set_controls_enabled(True)

    def _export_and_email_combined_report(self):
        if not self.failed_inventory_page: QMessageBox.warning(self, "Error",
                                                               "The failed inventory page is not linked. Cannot create a combined report."); return
        passed_df = self.current_inventory_df;
        failed_df = self.failed_inventory_page.current_inventory_df
        if passed_df.empty and failed_df.empty: QMessageBox.information(self, "Email Failed",
                                                                        "No data available in either report to email."); return
        settings = QSettings("MyCompany", "FGInventoryApp");
        email_config = {"sender_email": settings.value("email/sender_email", ""),
                        "sender_password": settings.value("email/sender_password", ""),
                        "smtp_server": settings.value("email/smtp_server", ""),
                        "smtp_port": settings.value("email/smtp_port", 0, type=int),
                        "recipient_email": settings.value("email/recipient_email", "")}
        if not all([email_config["sender_email"], email_config["sender_password"], email_config["smtp_server"],
                    email_config["smtp_port"], email_config["recipient_email"]]): QMessageBox.warning(self,
                                                                                                      "Configuration Needed",
                                                                                                      "Email settings are incomplete. Please use the 'Settings' button to configure them first."); self._open_settings_dialog(); return
        try:
            attachments = {};
            if not passed_df.empty:
                passed_export_df = passed_df.rename(
                    columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                             'location': 'LOCATION', 'bag_box_number': 'BAG/BOX_NUMBER'});
                final_cols_passed = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
                df_mb_passed = passed_export_df[passed_export_df['fg_type'] == 'MB'][final_cols_passed].copy();
                df_dc_passed = passed_export_df[passed_export_df['fg_type'] == 'DC'][final_cols_passed].copy();
                passed_dict = {};
                if not df_mb_passed.empty: passed_dict['FG MB PASSED'] = df_mb_passed
                if not df_dc_passed.empty: passed_dict['FG DC PASSED'] = df_dc_passed
                if passed_dict: attachments[f"FG_Inventory_Report_PASSED_{datetime.now():%Y-%m-%d}.xlsx"] = passed_dict
            if not failed_df.empty:
                failed_export_df = failed_df.rename(
                    columns={'product_code': 'PRODUCT_CODE', 'lot_number': 'LOT_NUMBER', 'current_balance': 'QTY',
                             'bag_box_number': 'BAG/BOX_NUMBER', 'location': 'LOCATION'});
                final_cols_failed = ['PRODUCT_CODE', 'LOT_NUMBER', 'BAG/BOX_NUMBER', 'QTY', 'LOCATION'];
                df_mb_failed = failed_export_df[failed_export_df['fg_type'] == 'MB'][final_cols_failed].copy();
                df_dc_failed = failed_export_df[failed_export_df['fg_type'] == 'DC'][final_cols_failed].copy();
                failed_dict = {};
                if not df_mb_failed.empty: failed_dict['FG FAILED MB'] = df_mb_failed
                if not df_dc_failed.empty: failed_dict['FG FAILED DC'] = df_dc_failed
                if failed_dict: attachments[f"FG_Inventory_Report_FAILED_{datetime.now():%Y-%m-%d}.xlsx"] = failed_dict
            if not attachments: QMessageBox.warning(self, "Email Failed",
                                                    "No data to include in the email after filtering."); return
            as_of_date_str = self.date_picker.date().toString("yyyy-MM-dd");
            email_config['subject'] = f"COMBINED Inventory Report (Passed & Failed) as of {as_of_date_str}";
            self.set_controls_enabled(False)
            if self.failed_inventory_page: self.failed_inventory_page.set_controls_enabled(False)
            QApplication.processEvents();
            send_email_with_excel(attachments, email_config);
            self.log_audit_trail("EMAIL_COMBINED_INVENTORY",
                                 f"User emailed combined inventory report to '{email_config['recipient_email']}'.");
            QMessageBox.information(self, "Email Sent",
                                    f"The combined report has been successfully sent to {email_config['recipient_email']}.")
        except Exception as e:
            show_error_message(self, "Email Error", "Failed to send combined email.",
                               f"Error details:\n\n{traceback.format_exc()}")
        finally:
            self.set_controls_enabled(True)
            if self.failed_inventory_page: self.failed_inventory_page.set_controls_enabled(True)

    def _handle_endorsement_summary_request(self):
        dialog = DateRangeDialog(self)
        if dialog.exec():
            start_date, end_date = dialog.get_dates()
            if start_date > end_date: show_error_message(self, "Invalid Date Range",
                                                         "The start date cannot be after the end date."); return
            start_date_str = start_date.toString(Qt.DateFormat.ISODate);
            end_date_str = end_date.toString(Qt.DateFormat.ISODate)
            self.set_controls_enabled(False);
            self.log_audit_trail("ENDORSEMENT_SUMMARY_START",
                                 f"User initiated endorsement summary for {start_date_str} to {end_date_str}")
            self.endorsement_summary_thread = QThread();
            self.endorsement_summary_worker = EndorsementSummaryWorker(self.engine, start_date_str, end_date_str);
            self.endorsement_summary_worker.moveToThread(self.endorsement_summary_thread)
            self.endorsement_summary_thread.started.connect(self.endorsement_summary_worker.run);
            self.endorsement_summary_worker.finished.connect(self._on_endorsement_summary_finished);
            self.endorsement_summary_worker.error.connect(self._on_calculation_error)
            self.endorsement_summary_worker.finished.connect(self.endorsement_summary_thread.quit);
            self.endorsement_summary_thread.finished.connect(self.endorsement_summary_worker.deleteLater);
            self.endorsement_summary_thread.finished.connect(self.endorsement_summary_thread.deleteLater);
            self.endorsement_summary_thread.start()

    def _on_endorsement_summary_finished(self, df: pd.DataFrame):
        self.set_controls_enabled(True)
        if df.empty: QMessageBox.information(self, "Report Complete",
                                             "No net endorsements found for the selected date range."); return
        default_filename = f"Endorsement_Summary_Report_{datetime.now():%Y-%m-%d}.xlsx";
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Endorsement Summary", default_filename,
                                                  "Excel Files (*.xlsx)")
        if filepath:
            try:
                # Define the target Quantity header
                QTY_COL_HEADER = 'TOTAL_QTY (Net: Passed - Failed)'

                # 1. Rename columns precisely as requested
                export_df = df.rename(
                    columns={'product_code': 'PRODUCT_CODE',
                             'total_qty': QTY_COL_HEADER,
                             'aggregated_remarks': 'REMARKS'})

                # 2. Select and enforce the required column order: Product, Qty, Remarks
                export_df = export_df[['PRODUCT_CODE', QTY_COL_HEADER, 'REMARKS']]

                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='Endorsement Summary');

                    # Ensure formatting is applied, searching for the key "QTY" in the header
                    self._format_excel_sheet(writer.sheets['Endorsement Summary'], export_df)

                filename = os.path.basename(filepath);
                self.log_audit_trail("ENDORSEMENT_SUMMARY_EXPORT",
                                     f"User exported endorsement summary to '{filename}'.");
                QMessageBox.information(self, "Export Successful",
                                        f"Endorsement Summary Report exported to:\n{filename}")
            except Exception:
                show_error_message(self, "Export Error", "Failed to save the endorsement summary file.",
                                   traceback.format_exc())


if __name__ == '__main__':
    app = QApplication(sys.argv)

    # --- Database Connection Setup (Adjust these parameters as needed) ---
    DB_USER = "postgres"
    DB_PASS = "mbpi"
    DB_HOST = "192.168.1.13"
    DB_PORT = "5432"
    DB_NAME = "dbfg"
    POSTGRES_URL = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

    db_engine = None
    try:
        db_engine = create_engine(POSTGRES_URL)
        # Attempt a quick connection test
        with db_engine.connect() as conn:
            conn.execute(text("SELECT 1"))
            print("Successfully connected to the database.")
    except Exception as e:
        QMessageBox.critical(None, "Database Error",
                             f"Could not connect to the database.\nPlease check your connection settings.\n\nError: {e}");
        sys.exit(1)


    def mock_log_audit_trail(action, details):
        print(f"[AUDIT LOG] User: StandaloneUser, Action: {action}, Details: {details}")


    class MockFailedInventoryPage(QWidget):
        def __init__(self):
            super().__init__()
            # Mock data for demonstration purposes, simulating failed inventory
            data = {'product_code': ['FAIL-01'], 'lot_number': ['FAIL-LOT'], 'current_balance': [-50.0],
                    'bag_box_number': [5], 'location': ['QUARANTINE'], 'fg_type': ['MB']}
            self.current_inventory_df = pd.DataFrame(data)

        def set_controls_enabled(self, enabled: bool):
            # Mock function to simulate control enabling/disabling
            print(f"MockFailedInventoryPage controls enabled: {enabled}")


    main_window = QWidget();
    main_window.setWindowTitle("Inventory Management System");
    main_layout = QVBoxLayout(main_window)

    # Instantiate the main inventory page
    good_inventory_page = GoodInventoryPage(engine=db_engine, username="StandaloneUser",
                                            log_audit_trail_func=mock_log_audit_trail)

    # Link the mock failed inventory page for the combined email feature
    good_inventory_page.failed_inventory_page = MockFailedInventoryPage()

    main_layout.addWidget(good_inventory_page);
    main_window.resize(1280, 800);
    main_window.show()

    sys.exit(app.exec())